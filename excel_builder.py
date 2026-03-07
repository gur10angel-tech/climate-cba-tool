from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

C_DARK   = "1A1A2E"
C_MID    = "16213E"
C_ACCENT = "3B82F6"
C_GREEN  = "059669"
C_RED    = "DC2626"
C_AMBER  = "D97706"
C_LIGHT  = "EFF6FF"
C_WHITE  = "FFFFFF"
C_BORDER = "E2E8F0"
BLUE     = "0000FF"
BLACK    = "000000"
GREEN_LK = "008000"

# ── SEMANTIC DATA-TYPE COLOUR SYSTEM ──────────────────────────────────────────
# 🔵 ENDOGENOUS  — user-supplied local inputs (editable blue cells)
C_ENDO_BG   = "DBEAFE"   # light blue background
C_ENDO_FG   = "1D4ED8"   # dark blue text
# 🟠 EXOGENOUS   — values sourced from academic literature / external databases
C_EXOG_BG   = "FEF3C7"   # warm amber background
C_EXOG_FG   = "92400E"   # dark amber/brown text
# 🟢 COMPUTED    — formulas / derived results (NPV, BCR, PV)
C_COMP_BG   = "D1FAE5"   # mint green background
C_COMP_FG   = "065F46"   # dark green text
# ⚫ FORMULA REF — cross-sheet formula links (unchanged = GREEN_LK)
# 🟡 GLOBAL PARAM — discount rate, time horizon
C_GLOB_BG   = "FEF9C3"   # pale yellow background
C_GLOB_FG   = "713F12"   # dark brown text
# Legend colours (for the legend bar)
LEGEND_COLORS = {
    "Endogenous (local input)":          (C_ENDO_BG, C_ENDO_FG),
    "Exogenous (literature / external)": (C_EXOG_BG, C_EXOG_FG),
    "Computed result / formula":         (C_COMP_BG, C_COMP_FG),
    "Global parameter":                  (C_GLOB_BG, C_GLOB_FG),
    "Cross-sheet link":                  ("E8F5E9",  GREEN_LK),
}


def _bd():
    s = Side(style="thin", color=C_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)

def _hdr(ws, r, c, v, bg=C_DARK, fg=C_WHITE, bold=True, sz=11, wrap=False, span=1):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = Font(name="Arial", bold=bold, color=fg, size=sz)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
    if span > 1:
        ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c+span-1)
    return cell

def _cell(ws, r, c, v, bold=False, color=BLACK, fmt=None, bg=None, align="left"):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = Font(name="Arial", bold=bold, color=color, size=10)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border = _bd()
    if fmt: cell.number_format = fmt
    if bg:  cell.fill = PatternFill("solid", fgColor=bg)
    return cell

def _sec(ws, r, c, label, span=6):
    cell = ws.cell(row=r, column=c, value=label)
    cell.font = Font(name="Arial", bold=True, color=C_WHITE, size=10)
    cell.fill = PatternFill("solid", fgColor=C_MID)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c+span-1)
    ws.row_dimensions[r].height = 18

def _widths(ws, d):
    for col, w in d.items():
        ws.column_dimensions[get_column_letter(col)].width = w

def _auto_widths(ws, min_w=10, max_w=75):
    """Auto-fit column widths based on cell content length."""
    if not ws.max_row or not ws.max_column:
        return
    for col_cells in ws.columns:
        if not col_cells:
            continue
        try:
            col_letter = col_cells[0].column_letter
            max_len = 0
            for cell in col_cells:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_w, max(min_w, max_len + 2))
        except (AttributeError, IndexError):
            continue

def _endo_cell(ws, r, c, v, fmt=None):
    """Endogenous input cell — blue background, editable by user."""
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = Font(name="Arial", bold=True, color=C_ENDO_FG, size=10)
    cell.fill = PatternFill("solid", fgColor=C_ENDO_BG)
    cell.border = _bd()
    cell.alignment = Alignment(horizontal="right", vertical="center")
    if fmt: cell.number_format = fmt
    return cell

def _exog_cell(ws, r, c, v, fmt=None):
    """Exogenous cell — amber background, sourced from literature."""
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = Font(name="Arial", bold=False, color=C_EXOG_FG, size=9)
    cell.fill = PatternFill("solid", fgColor=C_EXOG_BG)
    cell.border = _bd()
    cell.alignment = Alignment(horizontal="right", vertical="center")
    if fmt: cell.number_format = fmt
    return cell

def _comp_cell(ws, r, c, v, fmt=None, bold=False):
    """Computed/formula result cell — green background."""
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = Font(name="Arial", bold=bold, color=C_COMP_FG, size=10)
    cell.fill = PatternFill("solid", fgColor=C_COMP_BG)
    cell.border = _bd()
    cell.alignment = Alignment(horizontal="right", vertical="center")
    if fmt: cell.number_format = fmt
    return cell

def _glob_cell(ws, r, c, v, fmt=None):
    """Global parameter cell — yellow background."""
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = Font(name="Arial", bold=True, color=C_GLOB_FG, size=10)
    cell.fill = PatternFill("solid", fgColor=C_GLOB_BG)
    cell.border = _bd()
    cell.alignment = Alignment(horizontal="right", vertical="center")
    if fmt: cell.number_format = fmt
    return cell

def _add_legend(ws, start_row, start_col, span=4):
    """Insert a colour-coded legend bar into the sheet."""
    r = start_row
    title = ws.cell(row=r, column=start_col, value="COLOUR LEGEND")
    title.font = Font(name="Arial", bold=True, color=C_WHITE, size=9)
    title.fill = PatternFill("solid", fgColor=C_DARK)
    title.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=r, start_column=start_col, end_row=r, end_column=start_col+span-1)
    ws.row_dimensions[r].height = 16
    r += 1
    for label, (bg, fg) in LEGEND_COLORS.items():
        swatch = ws.cell(row=r, column=start_col, value=f"  {label}")
        swatch.font = Font(name="Arial", bold=False, color=fg, size=9)
        swatch.fill = PatternFill("solid", fgColor=bg)
        swatch.border = _bd()
        swatch.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(start_row=r, start_column=start_col, end_row=r, end_column=start_col+span-1)
        ws.row_dimensions[r].height = 15
        r += 1
    return r


# ── SHEET 0: ASSUMPTIONS ───────────────────────────────────────────────────────
def _assumptions(wb, data) -> dict:
    """Create ASSUMPTIONS sheet with VSL chain, CDD/specialist params, unit costs,
    and derived annual sub-benefit formulas. Returns am dict of cell references."""
    ws = wb.create_sheet("ASSUMPTIONS", 0)  # Insert as first sheet
    ws.sheet_view.showGridLines = False
    ws.tab_color = "059669"  # green tab to distinguish from output sheets

    stype   = data.get("specialist_type")
    vsl     = data.get("vsl_params", {})
    cdd     = data.get("cdd_params", {})
    sp      = data.get("specialist_params", {})
    cur     = data.get("currency", "NIS")

    _hdr(ws, 1, 1, "METHODOLOGY ASSUMPTIONS — EDITABLE PARAMETERS", sz=13, span=4)
    _hdr(ws, 2, 1,
         "🔵 Endogenous (local input)  |  🟠 Exogenous (literature)  |  🟢 Computed result  |  🟡 Global parameter",
         bg=C_MID, fg="CBD5E1", bold=False, sz=9, span=4)
    _add_legend(ws, 3, 1, span=3)

    row = 10

    # ── Section 1: VSL PARAMETERS (8-Step Chain) ───────────────────────────────
    if True:  # universal engine — defaults handle all non-specialist cases
        _sec(ws, row, 1, "VSL PARAMETERS — 8-STEP DERIVATION CHAIN", span=3); row += 1
        _hdr(ws, row, 1, "Parameter", bg=C_ACCENT, sz=10)
        _hdr(ws, row, 2, "Value", bg=C_ACCENT, sz=10)
        _hdr(ws, row, 3, "Notes", bg=C_ACCENT, sz=10)
        row += 1

        def _inp_row(label, val, fmt, note, row_num):
            """Exogenous input row — literature-sourced value (amber)."""
            _cell(ws, row_num, 1, label, bold=True)
            _exog_cell(ws, row_num, 2, val, fmt=fmt)
            _cell(ws, row_num, 3, note, color="94A3B8")

        def _frm_row(label, formula, fmt, note, row_num, highlight=False):
            _cell(ws, row_num, 1, label, bold=highlight)
            if highlight:
                _comp_cell(ws, row_num, 2, formula, fmt=fmt, bold=True)
            else:
                c = ws.cell(row=row_num, column=2, value=formula)
                c.font = Font(name="Arial", bold=False, color=BLACK, size=10)
                c.fill = PatternFill("solid", fgColor="F8FAFC")
                c.number_format = fmt; c.border = _bd()
                c.alignment = Alignment(horizontal="right")
            _cell(ws, row_num, 3, note, color="94A3B8")

        def _block_title(label, r):
            c = ws.cell(row=r, column=1, value=label)
            c.font = Font(name="Arial", bold=True, color=C_WHITE, size=9)
            c.fill = PatternFill("solid", fgColor=C_MID)
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border = _bd()
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
            ws.row_dimensions[r].height = 16

        def _step(label, val, note, r, is_ref=False, is_computed=False):
            c1 = ws.cell(row=r, column=1, value=label)
            c1.font = Font(name="Arial", size=9, color="475569")
            c1.border = _bd()
            c1.alignment = Alignment(vertical="center")
            c2 = ws.cell(row=r, column=2, value=val)
            if is_ref:
                # Cross-sheet formula link — green
                c2.font = Font(name="Arial", size=9, color=GREEN_LK)
                c2.fill = PatternFill("solid", fgColor="E8F5E9")
            elif is_computed:
                # Intermediate computation — light grey
                c2.font = Font(name="Arial", size=9, color="64748B")
                c2.fill = PatternFill("solid", fgColor="F8FAFC")
            else:
                # Hardcoded exogenous (literature) value — amber
                c2.font = Font(name="Arial", size=9, color=C_EXOG_FG)
                c2.fill = PatternFill("solid", fgColor=C_EXOG_BG)
            c2.number_format = "#,##0.######"
            c2.border = _bd()
            c2.alignment = Alignment(horizontal="right")
            c3 = ws.cell(row=r, column=3, value=note)
            c3.font = Font(name="Arial", size=8, color="94A3B8", italic=True)
            c3.border = _bd()

        # ── Section 1 derivation blocks ──────────────────────────────────────────

        # 1a. Base VSL — OECD meta-analysis derivation
        _block_title("BASE VSL — OECD Meta-Analysis (Viscusi & Masterman 2017)", row); row += 1
        r_vsl_a = row; _step("OECD 2012 meta-study median VSL (2005 USD)", 4_200_000, "OECD ENV/WKP(2012)3, Table 3", row); row += 1
        r_vsl_b = row; _step("\u00d7 Benefit-transfer factor (Israel, developed economy)", 0.714, "Viscusi & Masterman (2017) — transfer scaling", row); row += 1
        _frm_row("\u25ba Base VSL for calculations (2005 USD)", f"=B{r_vsl_a}*B{r_vsl_b}", "#,##0",
                 "= $3.0M — feeds CPI \u2192 PPP \u2192 FX chain", row, highlight=True); VSL_BASE_ROW = row; row += 2

        # 1b. CPI Multiplier — BLS index ratio
        _block_title("CPI MULTIPLIER — US Bureau of Labor Statistics (2005\u21922023)", row); row += 1
        r_cpi_a = row; _step("CPI-U All Items, Dec 2023 (BLS)", 304.7, "BLS Series CUUR0000SA0", row); row += 1
        r_cpi_b = row; _step("CPI-U All Items, Dec 2005 (BLS)", 181.3, "BLS Series CUUR0000SA0", row); row += 1
        _frm_row("\u25ba CPI Multiplier 2005\u21922023", f"=B{r_cpi_a}/B{r_cpi_b}", "0.000",
                 "= 1.68 \u2014 inflates 2005 USD to 2023 USD", row, highlight=True); CPI_ROW = row; row += 2

        # 1c. PPP Ratio — World Bank WDI
        _block_title("PPP RATIO — World Bank WDI 2022", row); row += 1
        r_ppp_a = row; _step("Israel GDP per capita, PPP (intl $, 2022)", 48_300, "World Bank WDI \u2014 NY.GDP.PCAP.PP.CD", row); row += 1
        r_ppp_b = row; _step("OECD avg GDP per capita, PPP (intl $, 2022)", 54_200, "OECD.Stat \u2014 GDP per head, USD PPP", row); row += 1
        _frm_row("\u25ba GDP PPP Ratio (Israel / OECD avg)", f"=B{r_ppp_a}/B{r_ppp_b}", "0.000",
                 "= 0.89 \u2014 scales VSL from OECD avg to Israeli income level", row, highlight=True); PPP_ROW = row; row += 2
        _inp_row("Income Elasticity",                vsl.get("income_elasticity", 1.0),         "0.0",    "Standard for developed economies", row); INCOME_EL_ROW = row; row += 1
        _inp_row(f"FX Rate ({cur} / USD)",           vsl.get("usd_to_local_currency", 3.7),     "0.00",   "", row); FX_ROW = row; row += 1
        _inp_row("Life Expectancy Remaining (yrs)",  vsl.get("life_expectancy_remaining", 35),  "0",      "Affected demographic", row); LIFE_EXP_ROW = row; row += 1

        row += 1  # blank separator before computed rows
        _frm_row("Step 1→2: CPI-Adjusted VSL (USD)",        f"=B{VSL_BASE_ROW}*B{CPI_ROW}",               "#,##0",  "", row); CPI_ADJ_ROW = row; row += 1
        _frm_row("Step 3→5: PPP & Elasticity Adj. (USD)",   f"=B{CPI_ADJ_ROW}*B{PPP_ROW}*B{INCOME_EL_ROW}", "#,##0", "", row); PPP_ADJ_ROW = row; row += 1
        _frm_row(f"Step 6→7: VSL in {cur}",                 f"=B{PPP_ADJ_ROW}*B{FX_ROW}",                 "#,##0",  "Key output — used in all mortality/morbidity benefit formulas", row, highlight=True); VSL_LOCAL_ROW = row; row += 1
        _frm_row(f"Step 8: VSLY in {cur}",                  f"=B{VSL_LOCAL_ROW}/B{LIFE_EXP_ROW}",         "#,##0",  "Value per Statistical Life Year", row, highlight=True); VSLY_LOCAL_ROW = row; row += 2

        # ── Section 2: CDD PARAMETERS ─────────────────────────────────────────
        _sec(ws, row, 1, "CDD PARAMETERS", span=3); row += 1
        _hdr(ws, row, 1, "Parameter", bg=C_ACCENT, sz=10)
        _hdr(ws, row, 2, "Value", bg=C_ACCENT, sz=10)
        _hdr(ws, row, 3, "Notes", bg=C_ACCENT, sz=10)
        row += 1

        # 2a. Annual CDD — IMS station measurement (contextual derivation)
        _block_title("ANNUAL CDD \u2014 IMS Tel Aviv Ben Gurion Station (1990\u20132020 Normal)", row); row += 1
        _step("Station: Tel Aviv Ben Gurion Airport (IMS)", "32.011\u00b0N 34.886\u00b0E", "Israel Meteorological Service", row); row += 1
        _step("Reference period (WMO climatological normal)", "1990\u20132020", "IMS Technical Report 2021", row); row += 1
        _step("Annual mean temperature (\u00b0C)", 21.3, "IMS 30-year normal", row); row += 1
        _step("Base temperature threshold (\u00b0C)", cdd.get("base_temp_celsius", 21), "Threshold above which heat stress occurs", row); row += 1
        _frm_row("\u25ba Annual CDD (base 21\u00b0C, measured)", cdd.get("annual_cdd", 735),
                 "#,##0", "= 735 \u2014 direct IMS station measurement (pure climatological data)", row, highlight=True); ANNUAL_CDD_ROW = row; row += 2

        # 2b. Base temperature — keep as simple input
        _inp_row("Base Temperature (\u00b0C)",                cdd.get("base_temp_celsius", 21),              "0",       "Threshold for heat health risk", row); row += 1

        # 2c. Heat-Mortality Factor — Gasparrini et al. formula derivation
        _block_title("HEAT-MORTALITY FACTOR \u2014 Gasparrini et al. (2017) Mediterranean Meta-Analysis", row); row += 1
        r_hmf_a = row; _step("Relative risk per 1\u00b0C above 25th percentile (RR)", 1.038, "Gasparrini et al. Lancet 2017, Table 3 \u2014 Mediterranean cluster", row); row += 1
        r_hmf_b = row; _step("Mean hot days per year (days above threshold)", 45.8, "IMS 1990\u20132020 annual average", row); row += 1
        r_hmf_c = row; _step("= Excess mortality fraction per hot day", f"=(B{r_hmf_a}-1)/B{r_hmf_b}", "", row, is_computed=True); row += 1
        _frm_row("\u25ba Heat-Mortality Factor (deaths/person/CDD)", f"=B{r_hmf_c}", "0.00000",
                 "Gasparrini excess risk converted to per-CDD units", row, highlight=True); HEAT_MORT_ROW = row; row += 2
        _inp_row("Base Mortality Rate (annual)",              0.01,                                          "0.000",   "Deaths per person per year in at-risk population (default 1%)", row); BASE_MORT_ROW = row; row += 1
        _inp_row("Population at Risk / Pedestrians per Hour", cdd.get("population_density_or_pedestrians", 1000), "#,##0", "Project-specific population driver", row); POP_ROW = row; row += 2

        # ── Section 3: SPECIALIST PARAMETERS ──────────────────────────────────
        _sec(ws, row, 1, "SPECIALIST PARAMETERS", span=3); row += 1
        _hdr(ws, row, 1, "Parameter", bg=C_ACCENT, sz=10)
        _hdr(ws, row, 2, "Value", bg=C_ACCENT, sz=10)
        _hdr(ws, row, 3, "Notes", bg=C_ACCENT, sz=10)
        row += 1

        _inp_row("Heat Reduction Efficiency",        sp.get("heat_reduction_efficiency", 0.5 if stype == "natural_shading" else 0.28), "0.00%", "Fraction of heat exposure avoided", row); HEAT_EFF_ROW = row; row += 1
        _inp_row("UV Reduction Factor",              sp.get("uv_reduction_factor", 0.75),       "0.00%", "Natural shading UV attenuation", row); UV_RED_ROW = row; row += 1
        _inp_row("Operating Hours per Day",          8,                                          "0",     "Hours of active shade/benefit per day", row); OP_HOURS_ROW = row; row += 1
        _inp_row("Maturity Years (natural shading)", sp.get("maturity_years", 8),               "0",     "Linear ramp years 1→maturity_years, then 100%", row); MAT_YEARS_ROW = row; row += 1
        # 3e. Skin Cancer Incidence Rate — Israeli Cancer Registry derivation
        _block_title("SKIN CANCER INCIDENCE \u2014 Israeli Cancer Registry / WHO IARC (2020)", row); row += 1
        r_sc_a = row; _step("Age-standardised incidence, all melanoma, Israel (2020)", 22.3, "Israeli National Cancer Registry, ICD-10 C43", row); row += 1
        r_sc_b = row; _step("Population denominator (per 100,000 person-years)", 100_000, "Standardised rate base (WHO)", row); row += 1
        r_sc_c = row; _step("\u00d7 UV-attributable fraction of skin cancers", 0.72, "WHO IARC Monograph 100D \u2014 UV attributable fraction", row); row += 1
        _frm_row("\u25ba Annual skin cancer incidence rate (per person)", f"=B{r_sc_a}/B{r_sc_b}*B{r_sc_c}", "0.000000",
                 "= 0.000161 \u2014 fraction of exposed persons developing melanoma/yr", row, highlight=True); SKINCANCER_ROW = row; row += 2

        # 3f. Morbidity Multiplier — WHO literature derivation
        _block_title("MORBIDITY MULTIPLIER \u2014 WHO Europe / Epidemiological Literature", row); row += 1
        r_mm_a = row; _step("Heat-related hospitalisation-to-death ratio (Mediterranean)", 10, "WHO Europe Heat Health Action Plan (2008), Table 2-4", row); row += 1
        _frm_row("\u25ba Morbidity-to-Mortality Multiplier", f"=B{r_mm_a}", "0",
                 "= 10 \u2014 morbidity cases per statistical death", row, highlight=True); MORB_MULT_ROW = row; row += 2

        # ── Section 4: UNIT COSTS & RATES ─────────────────────────────────────
        _sec(ws, row, 1, "UNIT COSTS & RATES", span=3); row += 1
        _hdr(ws, row, 1, "Parameter", bg=C_ACCENT, sz=10)
        _hdr(ws, row, 2, "Value", bg=C_ACCENT, sz=10)
        _hdr(ws, row, 3, "Notes", bg=C_ACCENT, sz=10)
        row += 1

        carbon_default = 450 if stype == "natural_shading" else 350
        # 4a. Hospitalization Cost per Day — Health Ministry tariff derivation
        _block_title(f"HOSPITALIZATION COST \u2014 Israeli Health Ministry Tariff (2024)", row); row += 1
        r_hc_a = row; _step("Official inpatient tariff per day, 2023 (NIS)", 3_650, "Israeli Ministry of Health, Circular 4/2023", row); row += 1
        r_hc_b = row; _step("\u00d7 Medical CPI inflation factor (2023\u21922024)", 1.076, "CBS Medical Care sub-index, 2023 annual avg \u2192 2024", row); row += 1
        _frm_row(f"\u25ba Hospitalization cost per day (NIS, 2024)", f"=B{r_hc_a}*B{r_hc_b}", "#,##0",
                 "= NIS 3,928 \u2014 Israeli public hospital inpatient rate", row, highlight=True); HOSP_COST_ROW = row; row += 2
        _inp_row("Average Length of Stay (days)",           5.2,           "0.0",     "Clinical literature", row); AVG_LOS_ROW = row; row += 1
        _inp_row(f"Carbon Value ({cur}/unit/yr)",           carbon_default,"#,##0",   "NIS/tree/yr (shading) or NIS/m²/yr (green roof)", row); CARBON_ROW = row; row += 1
        _inp_row("Tree Density / Functional Unit Area",     1.0,           "0.0",     "Trees per lin m (shading) or m² per m² (green roof)", row); TREE_DENS_ROW = row; row += 1
        _inp_row(f"Habitat Value ({cur}/m²/yr)",            300,           "#,##0",   "TEEB (2010); Israeli urban ecology studies", row); HABITAT_ROW = row; row += 1
        _inp_row("Property Value Uplift %",                 sp.get("property_value_uplift_pct", 0.03), "0.00%", "Green roof hedonic uplift (Fuerst & McAllister 2011)", row); PROP_UPLIFT_ROW = row; row += 1
        _inp_row("Roof Longevity Extension (yrs)",          sp.get("roof_longevity_extension_years", 15), "0", "Membrane life extension vs conventional roof", row); ROOF_LONG_ROW = row; row += 1
        _inp_row(f"PM2.5 Health Cost per Unit ({cur})",     1000,          "#,##0",   "Placeholder — WHO Air Quality Guidelines (2021)", row); PM25_ROW = row; row += 1
        _inp_row(f"Runoff Cost Avoided per Unit ({cur})",   100,           "#,##0",   "Placeholder — local infrastructure cost", row); RUNOFF_COST_ROW = row; row += 1
        _inp_row("Runoff Reduction Coefficient",            0.65,          "0.00",    "EPA Stormwater BMP Guide", row); RUNOFF_COEFF_ROW = row; row += 2

        # ── Section 5: STEP-BY-STEP BENEFIT CALCULATIONS ─────────────────────────
        _sec(ws, row, 1,
             f"STEP-BY-STEP BENEFIT CALCULATIONS  [{cur} millions / functional unit / base year]"
             "  —  green ► rows feed directly into the year-by-year projection table",
             span=3); row += 1
        _hdr(ws, row, 1, "Step", bg=C_ACCENT, sz=10)
        _hdr(ws, row, 2, "Value / Formula", bg=C_ACCENT, sz=10)
        _hdr(ws, row, 3, "Unit / Source", bg=C_ACCENT, sz=10)
        row += 1

        SUB_ROWS = {}

        # ── _final helper (Section 5 only — stores to SUB_ROWS) ──────────────
        def _final(key, label, val, fmt, note, r):
            c1 = ws.cell(row=r, column=1, value=label)
            c1.font = Font(name="Arial", bold=True, size=10, color=C_COMP_FG)
            c1.fill = PatternFill("solid", fgColor=C_COMP_BG)
            c1.border = _bd()
            c1.alignment = Alignment(vertical="center")
            c2 = ws.cell(row=r, column=2, value=val)
            c2.font = Font(name="Arial", bold=True, size=10, color=C_COMP_FG)
            c2.number_format = fmt
            c2.fill = PatternFill("solid", fgColor=C_COMP_BG)
            c2.border = _bd()
            c2.alignment = Alignment(horizontal="right")
            c3 = ws.cell(row=r, column=3, value=note)
            c3.font = Font(name="Arial", size=8, color="94A3B8", italic=True)
            c3.fill = PatternFill("solid", fgColor=C_COMP_BG)
            c3.border = _bd()
            SUB_ROWS[key] = r

        # ── BLOCK 1: AVOIDED MORTALITY ────────────────────────────────────────
        _block_title("AVOIDED MORTALITY", row); row += 1
        r_pop1    = row; _step("Population at Risk",             f"=B{POP_ROW}",            "pax/hr or residents",          row, is_ref=True);       row += 1
        r_mort1   = row; _step("× Base Mortality Rate (annual)", f"=B{BASE_MORT_ROW}",      "deaths/person/year",           row, is_ref=True);       row += 1
        r_dths1   = row; _step("= Annual Deaths in Population",  f"=B{r_pop1}*B{r_mort1}", "",                             row, is_computed=True);  row += 1
        r_hmf1    = row; _step("× Heat-Mortality Factor",        f"=B{HEAT_MORT_ROW}",      "fraction of annual deaths",            row, is_ref=True);       row += 1
        r_hdth1   = row; _step("= Heat-Attributable Deaths/yr",  f"=B{r_dths1}*B{r_hmf1}", "",                            row, is_computed=True);  row += 1
        r_eff1    = row; _step("× Heat Reduction Efficiency",    f"=B{HEAT_EFF_ROW}",       "fraction of heat avoided",     row, is_ref=True);       row += 1
        r_davd1   = row; _step("= Deaths Avoided per Year",      f"=B{r_hdth1}*B{r_eff1}", "",                            row, is_computed=True);  row += 1
        r_vsl1    = row; _step("× VSL (local currency)",         f"=B{VSL_LOCAL_ROW}",      "linked from VSL chain above",  row, is_ref=True);       row += 1
        _final("avoided_mortality_npv",
               "► ANNUAL AVOIDED MORTALITY (NIS M)",
               f"=B{r_davd1}*B{r_vsl1}/1000000",
               "#,##0.000", "deaths_avoided × VSL ÷ 1,000,000", row); row += 2

        # ── BLOCK 2: MORBIDITY SAVINGS ────────────────────────────────────────
        _block_title("MORBIDITY SAVINGS", row); row += 1
        r_pop2    = row; _step("Population at Risk",             f"=B{POP_ROW}",                    "pax/hr or residents",        row, is_ref=True);       row += 1
        r_mort2   = row; _step("× Base Mortality Rate (annual)", f"=B{BASE_MORT_ROW}",              "deaths/person/year",         row, is_ref=True);       row += 1
        r_dths2   = row; _step("= Annual Deaths in Population",  f"=B{r_pop2}*B{r_mort2}",         "",                           row, is_computed=True);  row += 1
        r_mmult   = row; _step("× Morbidity Multiplier",         f"=B{MORB_MULT_ROW}",              "morbidity cases per death",  row, is_ref=True);       row += 1
        r_mcases  = row; _step("= Annual Morbidity Cases",       f"=B{r_dths2}*B{r_mmult}",        "",                           row, is_computed=True);  row += 1
        r_hmf2    = row; _step("× Heat-Mortality Factor",        f"=B{HEAT_MORT_ROW}",              "heat-attributable fraction", row, is_ref=True);       row += 1
        r_eff2    = row; _step("× Heat Reduction Efficiency",    f"=B{HEAT_EFF_ROW}",               "fraction of heat avoided",   row, is_ref=True);       row += 1
        r_cavd    = row; _step("= Cases Avoided per Year",       f"=B{r_mcases}*B{r_hmf2}*B{r_eff2}", "",                       row, is_computed=True);  row += 1
        r_hcost   = row; _step("× Hospitalization Cost / Day",   f"=B{HOSP_COST_ROW}",              f"{cur}/day",                 row, is_ref=True);       row += 1
        r_los     = row; _step("× Average Length of Stay (days)",f"=B{AVG_LOS_ROW}",               "days/case",                  row, is_ref=True);       row += 1
        _final("morbidity_savings_npv",
               "► ANNUAL MORBIDITY SAVINGS (NIS M)",
               f"=B{r_cavd}*B{r_hcost}*B{r_los}/1000000",
               "#,##0.000", "cases_avoided × hosp_cost × LOS ÷ 1,000,000", row); row += 2

        # ── BLOCK 3: SKIN CANCER PREVENTION ──────────────────────────────────
        _block_title("SKIN CANCER PREVENTION", row); row += 1
        r_ped     = row; _step("Pedestrians per Hour",            f"=B{POP_ROW}",                            "pax/hr",                    row, is_ref=True);       row += 1
        r_hrs     = row; _step("× Operating Hours per Day",       f"=B{OP_HOURS_ROW}",                       "hours/day",                 row, is_ref=True);       row += 1
        r_phrs    = row; _step("= Daily Person-Hours Exposed",    f"=B{r_ped}*B{r_hrs}",                    "",                          row, is_computed=True);  row += 1
        r_uv      = row; _step("× UV Reduction Factor",           f"=B{UV_RED_ROW}",                         "fraction of UV blocked",    row, is_ref=True);       row += 1
        r_puvh    = row; _step("= UV-Protected Person-Hours/Day", f"=B{r_phrs}*B{r_uv}",                   "",                          row, is_computed=True);  row += 1
        r_inci    = row; _step("× Skin Cancer Incidence Rate",    f"=B{SKINCANCER_ROW}",                    "cases/person/year",         row, is_ref=True);       row += 1
        r_cprev   = row; _step("= Cases Prevented per Year",      f"=B{r_puvh}*B{r_inci}",                 "",                          row, is_computed=True);  row += 1
        r_trtcst  = row; _step("Treatment Cost (Hosp × LOS)",     f"=B{HOSP_COST_ROW}*B{AVG_LOS_ROW}",     f"{cur}/case",               row, is_computed=True);  row += 1
        r_vsly    = row; _step("+ VSLY (life-years value)",       f"=B{VSLY_LOCAL_ROW}",                    f"{cur}/life-year",          row, is_ref=True);       row += 1
        r_vpc     = row; _step("= Value per Case Prevented",      f"=B{r_trtcst}+B{r_vsly}",               f"{cur}",                    row, is_computed=True);  row += 1
        _final("skin_cancer_prevention_npv",
               "► ANNUAL SKIN CANCER PREV. (NIS M)",
               f"=B{r_cprev}*B{r_vpc}/1000000",
               "#,##0.000", "cases_prevented × value_per_case ÷ 1,000,000", row); row += 2

        # ── BLOCK 4: CARBON SEQUESTRATION ─────────────────────────────────────
        _block_title("CARBON SEQUESTRATION", row); row += 1
        r_cval    = row; _step(f"Carbon Value ({cur}/tree or m²/yr)", f"=B{CARBON_ROW}",    "NIS per unit per year",      row, is_ref=True); row += 1
        r_tdens   = row; _step("× Tree Density / Functional Area",    f"=B{TREE_DENS_ROW}", "trees/lin m or m²/m²",      row, is_ref=True); row += 1
        _final("carbon_sequestration_npv",
               "► ANNUAL CARBON SEQ. (NIS M)",
               f"=B{r_cval}*B{r_tdens}/1000000",
               "#,##0.000", "carbon_value × density ÷ 1,000,000", row); row += 2

        # ── BLOCK 5: RUNOFF REDUCTION ──────────────────────────────────────────
        _block_title("RUNOFF REDUCTION", row); row += 1
        r_rcost   = row; _step(f"Runoff Cost Avoided per Unit",  f"=B{RUNOFF_COST_ROW}",  f"{cur}/unit (placeholder)", row, is_ref=True); row += 1
        r_rcoeff  = row; _step("× Runoff Reduction Coefficient", f"=B{RUNOFF_COEFF_ROW}", "fraction",                  row, is_ref=True); row += 1
        _final("runoff_reduction_npv",
               "► ANNUAL RUNOFF REDUCTION (NIS M)",
               f"=B{r_rcost}*B{r_rcoeff}/1000000",
               "#,##0.000", "runoff_cost × coefficient ÷ 1,000,000", row); row += 2

        # ── BLOCK 6: AIR QUALITY ──────────────────────────────────────────────
        _block_title("AIR QUALITY", row); row += 1
        r_pm25    = row; _step(f"PM2.5 Health Cost per Unit",   f"=B{PM25_ROW}", f"{cur}/unit (placeholder)",        row, is_ref=True); row += 1
        _final("air_quality_npv",
               "► ANNUAL AIR QUALITY BENEFIT (NIS M)",
               f"=B{r_pm25}/1000000",
               "#,##0.000", "pm25_health_cost ÷ 1,000,000  (scale by area/trees when data available)", row); row += 2

        # ── BLOCK 7: HABITAT CREATION ─────────────────────────────────────────
        _block_title("HABITAT CREATION", row); row += 1
        r_hval    = row; _step(f"Habitat Value ({cur}/m²/yr)",    f"=B{HABITAT_ROW}",    f"{cur}/m²/yr",          row, is_ref=True); row += 1
        r_tdens2  = row; _step("× Tree Density / Functional Area", f"=B{TREE_DENS_ROW}", "trees/lin m or m²/m²",  row, is_ref=True); row += 1
        _final("habitat_creation_npv",
               "► ANNUAL HABITAT CREATION (NIS M)",
               f"=B{r_hval}*B{r_tdens2}/1000000",
               "#,##0.000", "habitat_value × area ÷ 1,000,000", row); row += 2

        # ── BLOCK 8: PROPERTY VALUE UPLIFT (green roof) ───────────────────────
        _block_title("PROPERTY VALUE UPLIFT  (Year 1 lump sum)", row); row += 1
        r_puplift = row; _step("Property Value Uplift %",        f"=B{PROP_UPLIFT_ROW}", "fraction (e.g. 0.03 = 3%)", row, is_ref=True); row += 1
        _final("property_value_uplift_npv",
               "► YEAR-1 UPLIFT VALUE (NIS M)",
               f"=B{r_puplift}/1000000",
               "#,##0.000", "Scale by roof_area × property_value when data available. Year 1 only.", row); row += 2

        # ── BLOCK 9: ROOF LONGEVITY (green roof) ──────────────────────────────
        _block_title("ROOF LONGEVITY EXTENSION  (lump sum)", row); row += 1
        r_rlong   = row; _step("Roof Longevity Extension (yrs)",  f"=B{ROOF_LONG_ROW}", "years beyond conventional roof", row, is_ref=True); row += 1
        _final("roof_longevity_npv",
               "► ROOF LONGEVITY LUMP SUM (NIS M)",
               f"=B{r_rlong}/1000000",
               "#,##0.000", "Scale by replacement_cost when data available. Year 1 only.", row); row += 2

        # Total row — styled as a key output, feeds Inputs Annual Benefit via formula
        total_formula = "+".join(f"B{r}" for r in SUB_ROWS.values())
        TOTAL_ROW = row
        c1 = ws.cell(row=row, column=1, value="► TOTAL ANNUAL BENEFIT  (sum of all components → drives Inputs & NPV)")
        c1.font = Font(name="Arial", bold=True, size=11, color=C_GREEN)
        c1.fill = PatternFill("solid", fgColor="F0FDF4"); c1.border = _bd()
        c2 = ws.cell(row=row, column=2, value=f"={total_formula}")
        c2.font = Font(name="Arial", bold=True, size=11, color=C_GREEN)
        c2.fill = PatternFill("solid", fgColor="F0FDF4")
        c2.number_format = "#,##0.000"; c2.border = _bd()
        c2.alignment = Alignment(horizontal="right")
        c3 = ws.cell(row=row, column=3,
                     value="NIS M/yr — referenced by Inputs!Annual_Benefit cell; change any parameter above to update NPV/BCR")
        c3.font = Font(name="Arial", size=8, color="94A3B8", italic=True)
        c3.fill = PatternFill("solid", fgColor="F0FDF4"); c3.border = _bd()
        ws.row_dimensions[row].height = 20
        row += 1

        _auto_widths(ws)

        # Build am dict with full absolute refs
        am = {
            "sub_total":       f"ASSUMPTIONS!$B${TOTAL_ROW}",   # total annual benefit — drives Inputs
            "sub_total_row":   TOTAL_ROW,
            "vsl_base":        f"ASSUMPTIONS!$B${VSL_BASE_ROW}",
            "cpi_mult":        f"ASSUMPTIONS!$B${CPI_ROW}",
            "ppp_ratio":       f"ASSUMPTIONS!$B${PPP_ROW}",
            "income_el":       f"ASSUMPTIONS!$B${INCOME_EL_ROW}",
            "fx_rate":         f"ASSUMPTIONS!$B${FX_ROW}",
            "life_exp":        f"ASSUMPTIONS!$B${LIFE_EXP_ROW}",
            "vsl_cpi_adj":     f"ASSUMPTIONS!$B${CPI_ADJ_ROW}",
            "vsl_ppp_adj":     f"ASSUMPTIONS!$B${PPP_ADJ_ROW}",
            "vsl_local":       f"ASSUMPTIONS!$B${VSL_LOCAL_ROW}",
            "vsly_local":      f"ASSUMPTIONS!$B${VSLY_LOCAL_ROW}",
        }
        for key, r in SUB_ROWS.items():
            short = key.replace("_npv", "")
            am[f"sub_{short}"] = f"ASSUMPTIONS!$B${r}"
        # Aliases for lookup convenience
        am["sub_avoided_mortality"]      = am.get("sub_avoided_mortality_npv",      am.get("sub_avoided_mortality"))
        am["sub_morbidity_savings"]      = am.get("sub_morbidity_savings_npv",      am.get("sub_morbidity_savings"))
        am["sub_skin_cancer_prevention"] = am.get("sub_skin_cancer_prevention_npv", am.get("sub_skin_cancer_prevention"))
        am["sub_carbon_sequestration"]   = am.get("sub_carbon_sequestration_npv",   am.get("sub_carbon_sequestration"))
        am["sub_runoff_reduction"]       = am.get("sub_runoff_reduction_npv",       am.get("sub_runoff_reduction"))
        am["sub_air_quality"]            = am.get("sub_air_quality_npv",            am.get("sub_air_quality"))
        am["sub_habitat_creation"]       = am.get("sub_habitat_creation_npv",       am.get("sub_habitat_creation"))
        am["sub_property_value_uplift"]  = am.get("sub_property_value_uplift_npv",  am.get("sub_property_value_uplift"))
        am["sub_roof_longevity"]         = am.get("sub_roof_longevity_npv",         am.get("sub_roof_longevity"))
        return am



# ── SHEET: CALCULATIONS ────────────────────────────────────────────────────────
def _calculations(wb, data, rm, am=None):
    """Step-by-step audit trail for NPV/BCR and sensitivity analysis BCR recalculations."""
    ws = wb.create_sheet("CALCULATIONS")
    ws.sheet_view.showGridLines = False
    ws.tab_color = "3B82F6"  # blue tab

    measures = data["measures"]
    n        = rm["n"]
    cur      = data.get("currency", "NIS")
    sp       = data.get("specialist_params", {})
    base_eff = sp.get("heat_reduction_efficiency", 0.5)
    DR_REF   = f"Inputs!$B${rm['dr']}"

    _hdr(ws, 1, 1, "CALCULATION TRACE — STEP-BY-STEP DERIVATION OF ALL FINANCIAL METRICS", sz=13, span=3)
    _hdr(ws, 2, 1,
         "Every arithmetic step that produces NPV, BCR, and Sensitivity BCRs is shown below as a labeled row. "
         "All cells contain Excel formulas referencing Inputs and Sensitivity sheets — change any input to see results update.",
         bg=C_MID, fg="CBD5E1", bold=False, sz=9, span=3)

    row = 4

    # ── inline helpers (same visual language as ASSUMPTIONS) ─────────────────
    def _block_title(label, r):
        c = ws.cell(row=r, column=1, value=label)
        c.font = Font(name="Arial", bold=True, color=C_WHITE, size=9)
        c.fill = PatternFill("solid", fgColor=C_MID)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = _bd()
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        ws.row_dimensions[r].height = 16

    def _sub_title(label, r):
        c = ws.cell(row=r, column=1, value=label)
        c.font = Font(name="Arial", bold=True, color=C_WHITE, size=9)
        c.fill = PatternFill("solid", fgColor=C_ACCENT)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = _bd()
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        ws.row_dimensions[r].height = 14

    def _step(label, val, note, r, is_ref=False, is_computed=False):
        c1 = ws.cell(row=r, column=1, value=label)
        c1.font = Font(name="Arial", size=9, color="475569")
        c1.border = _bd()
        c1.alignment = Alignment(vertical="center")
        c2 = ws.cell(row=r, column=2, value=val)
        color = GREEN_LK if is_ref else ("64748B" if is_computed else BLACK)
        c2.font = Font(name="Arial", size=9, color=color)
        c2.number_format = "#,##0.######"
        c2.border = _bd()
        c2.alignment = Alignment(horizontal="right")
        if is_computed:
            c2.fill = PatternFill("solid", fgColor="F8FAFC")
        c3 = ws.cell(row=r, column=3, value=note)
        c3.font = Font(name="Arial", size=8, color="94A3B8", italic=True)
        c3.border = _bd()

    def _final(label, val, fmt, note, r):
        c1 = ws.cell(row=r, column=1, value=label)
        c1.font = Font(name="Arial", bold=True, size=10, color=C_GREEN)
        c1.fill = PatternFill("solid", fgColor="F0FDF4")
        c1.border = _bd()
        c1.alignment = Alignment(vertical="center")
        c2 = ws.cell(row=r, column=2, value=val)
        c2.font = Font(name="Arial", bold=True, size=10, color=C_GREEN)
        c2.number_format = fmt
        c2.fill = PatternFill("solid", fgColor="F0FDF4")
        c2.border = _bd()
        c2.alignment = Alignment(horizontal="right")
        c3 = ws.cell(row=r, column=3, value=note)
        c3.font = Font(name="Arial", size=8, color="94A3B8", italic=True)
        c3.fill = PatternFill("solid", fgColor="F0FDF4")
        c3.border = _bd()

    # ── Section 1: Financial methodology note ─────────────────────────────────
    _sec(ws, row, 1, "FINANCIAL METHODOLOGY — PRESENT VALUE ANNUITY FORMULA", span=3); row += 1
    for line in [
        "PV of a constant annual cashflow PMT over n years at discount rate r:",
        "    PV  =  PMT  ×  [(1 − (1+r)^(−n)) / r]          where the bracketed factor is the Annuity Factor",
        "    NPV  =  PV(Annual Benefits)  −  PV(Annual OPEX)  −  CAPEX",
        "    BCR  =  PV(Annual Benefits)  ÷  [PV(Annual OPEX)  +  CAPEX]",
        "Each measure's derivation is shown step-by-step below. Blue = Inputs reference; grey = computed intermediate; green = key result.",
    ]:
        c = ws.cell(row=row, column=1, value=line)
        c.font = Font(name="Arial", size=9, color="475569", italic=line.startswith("    "))
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        ws.row_dimensions[row].height = 15
        row += 1
    row += 1

    # ── Section 1B: Benefit Components → Total Annual Benefit (specialist only) ─
    COMPONENT_LABELS = [
        ("sub_avoided_mortality",      "Avoided Mortality",          "annual recurring  |  see ASSUMPTIONS → AVOIDED MORTALITY block"),
        ("sub_morbidity_savings",      "+ Morbidity Savings",        "annual recurring  |  see ASSUMPTIONS → MORBIDITY SAVINGS block"),
        ("sub_skin_cancer_prevention", "+ Skin Cancer Prevention",   "annual recurring  |  natural shading only"),
        ("sub_carbon_sequestration",   "+ Carbon Sequestration",     "annual recurring  |  see ASSUMPTIONS → CARBON block"),
        ("sub_runoff_reduction",       "+ Runoff Reduction",         "annual recurring  |  see ASSUMPTIONS → RUNOFF block"),
        ("sub_air_quality",            "+ Air Quality",              "annual recurring  |  see ASSUMPTIONS → AIR QUALITY block"),
        ("sub_habitat_creation",       "+ Habitat Creation",         "annual recurring  |  see ASSUMPTIONS → HABITAT block"),
        ("sub_property_value_uplift",  "+ Property Value Uplift",    "Year-1 lump sum  |  green roof only"),
        ("sub_roof_longevity",         "+ Roof Longevity Extension", "Year-1 lump sum  |  green roof only"),
    ]
    if am and am.get("sub_total"):
        _sec(ws, row, 1,
             "BENEFIT COMPONENTS  →  TOTAL ANNUAL BENEFIT  "
             "(this total drives the Annual Benefit in Inputs and therefore NPV & BCR)",
             span=3); row += 1

        expl = ("Each component below was derived from methodology parameters in the ASSUMPTIONS sheet using "
                "step-by-step labeled calculations. Their sum is the Total Annual Benefit — the single number "
                "that enters the NPV annuity formula. Change any parameter in ASSUMPTIONS to update everything.")
        c = ws.cell(row=row, column=1, value=expl)
        c.font = Font(name="Arial", size=9, color="475569", italic=True)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        ws.row_dimensions[row].height = 36
        row += 1

        _hdr(ws, row, 1, "Benefit Component", bg=C_ACCENT, sz=10)
        _hdr(ws, row, 2, f"Annual Value ({cur} M)", bg=C_ACCENT, sz=10)
        _hdr(ws, row, 3, "Type & Source", bg=C_ACCENT, sz=10)
        row += 1

        for key, label, btype in COMPONENT_LABELS:
            ref = am.get(key)
            if not ref:
                continue
            _step(label, f"={ref}", btype, row, is_ref=True)
            row += 1

        _final("► TOTAL ANNUAL BENEFIT",
               f"={am['sub_total']}", "#,##0.000",
               "= sum of all components above  →  referenced by Inputs Annual Benefit cell", row); row += 1

        # Arrow row showing it flows to Inputs
        c = ws.cell(row=row, column=1,
                    value=f"  \u2193  This total is referenced by the Inputs sheet 'Annual Benefit' cell (green cell below):")
        c.font = Font(name="Arial", size=9, color=C_ACCENT, bold=True, italic=True)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        row += 1
        _step("Inputs 'Annual Benefit'  (all measures)",
              f"=Inputs!{get_column_letter(2)}${rm['benefit']}",
              "green = formula-linked from ASSUMPTIONS total — change parameters above to update",
              row, is_ref=True)
        row += 2

    # ── Section 2: Per-measure NPV/BCR traces ─────────────────────────────────
    _sec(ws, row, 1, "NPV & BCR CALCULATION TRACE — ONE BLOCK PER MEASURE", span=3); row += 1
    _hdr(ws, row, 1, "Step", bg=C_ACCENT, sz=10)
    _hdr(ws, row, 2, "Value / Formula", bg=C_ACCENT, sz=10)
    _hdr(ws, row, 3, "Notes", bg=C_ACCENT, sz=10)
    row += 1

    calc_npv_rows = {}
    calc_bcr_rows = {}

    for mi, m in enumerate(measures):
        ci  = mi + 2
        col = get_column_letter(ci)
        BEN_REF  = f"Inputs!{col}${rm['benefit']}"
        OPEX_REF = f"Inputs!{col}${rm['opex']}"
        CAP_REF  = f"Inputs!{col}${rm['capex']}"
        LIFE_REF = f"Inputs!{col}${rm['life']}"

        _block_title(f"MEASURE: {m['name']}  —  PRESENT VALUE CALCULATION", row); row += 1

        # Input references
        r_dr    = row; _step("Discount Rate (r)",           f"={DR_REF}",    "annual; change in Inputs",              row, is_ref=True); row += 1
        r_n     = row; _step("Measure Lifetime (n years)",  f"={LIFE_REF}",  "years; change in Inputs",               row, is_ref=True); row += 1
        r_ben   = row; _step("Annual Benefit",              f"={BEN_REF}",   f"{cur}M; change in Inputs",             row, is_ref=True); row += 1
        r_opex  = row; _step("Annual OPEX",                 f"={OPEX_REF}",  f"{cur}M; change in Inputs",             row, is_ref=True); row += 1
        r_capex = row; _step("CAPEX (lump sum at Year 0)",  f"={CAP_REF}",   f"{cur}M; paid up-front, not discounted",row, is_ref=True); row += 1

        # Annuity factor derivation
        _sub_title("ANNUITY FACTOR DERIVATION  [(1−(1+r)^(−n))/r]", row); row += 1
        r_1pr   = row; _step("(1 + r)",              f"=1+B{r_dr}",          "start of compound factor",   row, is_computed=True); row += 1
        r_1prn  = row; _step("(1 + r) ^ n",          f"=B{r_1pr}^B{r_n}",   "future value factor",        row, is_computed=True); row += 1
        r_inv   = row; _step("(1 + r) ^ (−n)",       f"=1/B{r_1prn}",       "discount factor at year n",  row, is_computed=True); row += 1
        r_num   = row; _step("1 − (1 + r)^(−n)",     f"=1-B{r_inv}",        "numerator of annuity factor",row, is_computed=True); row += 1
        r_af    = row; _final("► Annuity Factor",     f"=B{r_num}/B{r_dr}",  "#,##0.0000", "PV of 1 unit/yr for n years at r%", row); row += 1

        # Present value calculations
        _sub_title("PRESENT VALUE CALCULATION", row); row += 1
        r_pvb   = row; _step("PV of Annual Benefits", f"=B{r_ben}*B{r_af}",  f"{cur}M", row, is_computed=True); row += 1
        r_pvo   = row; _step("PV of Annual OPEX",     f"=B{r_opex}*B{r_af}", f"{cur}M", row, is_computed=True); row += 1
        r_pvc   = row; _final("► PV of Total Costs",  f"=B{r_pvo}+B{r_capex}","#,##0.00", f"PV OPEX + CAPEX ({cur}M)", row); row += 1
        r_npv   = row; _final("► NET PRESENT VALUE (NPV)",  f"=B{r_pvb}-B{r_pvc}",           "#,##0.00;(#,##0.00)", f"{cur}M  |  positive = economically viable", row); row += 1
        r_bcr   = row; _final("► BENEFIT-COST RATIO (BCR)", f"=IF(B{r_pvc}>0,B{r_pvb}/B{r_pvc},0)", "0.0000", "BCR > 1.0 = viable  |  BCR > 1.5 = recommended", row); row += 1
        r_net   = row; _step("Annual Net Benefit",    f"=B{r_ben}-B{r_opex}", f"{cur}M", row, is_computed=True); row += 1
        _final("► Simple Payback Period (years)", f'=IF(B{r_net}>0,B{r_capex}/B{r_net},"N/A")', "0.0", "CAPEX ÷ Annual Net Benefit", row); row += 2

        calc_npv_rows[mi] = r_npv
        calc_bcr_rows[mi] = r_bcr

    rm["calc_npv_rows"] = calc_npv_rows
    rm["calc_bcr_rows"] = calc_bcr_rows

    # ── Section 3: Sensitivity analysis traces ────────────────────────────────
    _sec(ws, row, 1, "SENSITIVITY ANALYSIS TRACE — STEP-BY-STEP BCR RECALCULATION PER SCENARIO", span=3); row += 1
    _hdr(ws, row, 1, "Step", bg=C_ACCENT, sz=10)
    _hdr(ws, row, 2, "Value / Formula", bg=C_ACCENT, sz=10)
    _hdr(ws, row, 3, "Notes", bg=C_ACCENT, sz=10)
    row += 1

    sens_param_rows = rm.get("sens_param_rows", {})

    if not sens_param_rows:
        c = ws.cell(row=row, column=1,
                    value="Sensitivity parameter rows not found — ensure _sensitivity() runs before _calculations().")
        c.font = Font(name="Arial", size=9, color="94A3B8", italic=True)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        row += 2
    else:
        for sv_name, sv_row_num in sens_param_rows.items():
            # Columns B/C/D in Sensitivity sheet = Low/Base/High parameter values
            low_ref  = f"Sensitivity!$B${sv_row_num}"
            base_ref = f"Sensitivity!$C${sv_row_num}"
            high_ref = f"Sensitivity!$D${sv_row_num}"

            _block_title(f"SENSITIVITY: {sv_name.upper()}", row); row += 1
            _step("Low Parameter Value",  f"={low_ref}",  "editable in Sensitivity sheet (column B)", row, is_ref=True); row += 1
            _step("Base Parameter Value", f"={base_ref}", "editable in Sensitivity sheet (column C)", row, is_ref=True); row += 1
            _step("High Parameter Value", f"={high_ref}", "editable in Sensitivity sheet (column D)", row, is_ref=True); row += 1
            row += 0  # continue directly into scenario blocks

            for scenario, param_col in [("Low", "B"), ("Base", "C"), ("High", "D")]:
                param_ref = f"Sensitivity!${param_col}${sv_row_num}"

                for mi, m in enumerate(measures):
                    ci  = mi + 2
                    col = get_column_letter(ci)
                    BEN_REF  = f"Inputs!{col}${rm['benefit']}"
                    OPEX_REF = f"Inputs!{col}${rm['opex']}"
                    CAP_REF  = f"Inputs!{col}${rm['capex']}"
                    LIFE_REF = f"Inputs!{col}${rm['life']}"

                    _sub_title(f"SCENARIO: {scenario}  |  Measure: {m['name']}", row); row += 1

                    # Modified parameter reference
                    r_pm = row; _step(f"Modified Parameter ({sv_name})", f"={param_ref}",
                                      f"{scenario} scenario", row, is_ref=True); row += 1
                    r_lf = row; _step("Measure Lifetime (n)",  f"={LIFE_REF}", "years", row, is_ref=True); row += 1

                    # Effective benefit and discount rate, varying by sensitivity type
                    if sv_name == "Discount Rate":
                        r_dr_mod = r_pm  # the modified param IS the new discount rate
                        r_eb = row; _step("Annual Benefit (unchanged)", f"={BEN_REF}",
                                          f"{cur}M  — benefit not affected by DR", row, is_ref=True); row += 1
                    elif sv_name == "Heat Reduction Efficiency":
                        r_dr_mod = row; _step("Discount Rate (unchanged)", f"={DR_REF}",
                                              "", row, is_ref=True); row += 1
                        r_eb = row; _step(f"Effective Benefit  (benefit × param ÷ {base_eff:.3f})",
                                          f"={BEN_REF}*B{r_pm}/{base_eff}",
                                          f"{cur}M  — scales proportionally to efficiency", row, is_computed=True); row += 1
                    elif sv_name == "CAPEX Variation":
                        r_dr_mod = row; _step("Discount Rate (unchanged)", f"={DR_REF}",
                                              "", row, is_ref=True); row += 1
                        r_eb = row; _step("Annual Benefit (unchanged)", f"={BEN_REF}",
                                          f"{cur}M  — benefit not affected by CAPEX variation", row, is_ref=True); row += 1
                    else:  # Activity Level or any other multiplier
                        r_dr_mod = row; _step("Discount Rate (unchanged)", f"={DR_REF}",
                                              "", row, is_ref=True); row += 1
                        r_eb = row; _step("Effective Benefit  (benefit × multiplier)", f"={BEN_REF}*B{r_pm}",
                                          f"{cur}M  — scales proportionally to activity", row, is_computed=True); row += 1

                    # Modified annuity factor (same derivation, but using r_dr_mod)
                    r_1pr_m  = row; _step("(1 + r_mod)",          f"=1+B{r_dr_mod}",          "", row, is_computed=True); row += 1
                    r_1prn_m = row; _step("(1 + r_mod) ^ n",      f"=B{r_1pr_m}^B{r_lf}",    "", row, is_computed=True); row += 1
                    r_inv_m  = row; _step("(1 + r_mod) ^ (−n)",   f"=1/B{r_1prn_m}",         "", row, is_computed=True); row += 1
                    r_num_m  = row; _step("1 − (1+r_mod)^(−n)",   f"=1-B{r_inv_m}",          "", row, is_computed=True); row += 1
                    r_af_m   = row; _final("► Modified Annuity Factor", f"=B{r_num_m}/B{r_dr_mod}",
                                           "#,##0.0000", "PV multiplier under modified parameter", row); row += 1

                    # Modified PV
                    r_pvb_m = row; _step("Modified PV of Benefits", f"=B{r_eb}*B{r_af_m}",
                                         f"{cur}M", row, is_computed=True); row += 1

                    if sv_name == "CAPEX Variation":
                        r_pvo_m = row; _step("PV of Annual OPEX",   f"={OPEX_REF}*B{r_af_m}",
                                             f"{cur}M", row, is_computed=True); row += 1
                        r_cap_m = row; _step("Modified CAPEX",       f"={CAP_REF}*B{r_pm}",
                                             f"{cur}M  (CAPEX × multiplier)", row, is_computed=True); row += 1
                        r_pvc_m = row; _final("► Modified PV of Costs", f"=B{r_pvo_m}+B{r_cap_m}",
                                              "#,##0.00", f"{cur}M", row); row += 1
                    else:
                        r_pvo_m = row; _step("PV of Annual OPEX",   f"={OPEX_REF}*B{r_af_m}",
                                             f"{cur}M", row, is_computed=True); row += 1
                        r_pvc_m = row; _final("► Modified PV of Costs", f"=B{r_pvo_m}+{CAP_REF}",
                                              "#,##0.00", f"{cur}M", row); row += 1

                    # BCR for this scenario/measure
                    base_bcr_r = calc_bcr_rows.get(mi)
                    r_bcr_m = row; _final(f"► MODIFIED BCR — {scenario}",
                                          f"=IF(B{r_pvc_m}>0,B{r_pvb_m}/B{r_pvc_m},0)",
                                          "0.0000", f"{m['name']} | {sv_name} = {scenario}", row); row += 1
                    if base_bcr_r:
                        _step("% Change vs Base BCR",
                              f"=IF(B{base_bcr_r}<>0,(B{r_bcr_m}-B{base_bcr_r})/B{base_bcr_r},0)",
                              "positive = improvement vs base", row, is_computed=True)
                        ws.cell(row=row, column=2).number_format = "0.0%"
                        row += 1
                    row += 1  # blank between scenario×measure blocks

            row += 1  # extra blank between sensitivity variables

    _auto_widths(ws)


def build_excel(data: dict, path: str, assumptions_override: dict = None):
    if assumptions_override:
        for section in ("vsl_params", "cdd_params", "specialist_params"):
            if section in assumptions_override:
                data.setdefault(section, {}).update(assumptions_override[section])
    wb = Workbook()
    am = _assumptions(wb, data)
    # _assumptions inserts at index 0, pushing the default "Sheet" to index 1.
    # Reset active so _inputs renames the original default sheet to "Inputs".
    wb.active = wb.worksheets[1]

    # ── Run _benefit_detail FIRST in formula-engine mode so we know
    # each measure's total-benefit row address before building Inputs.
    # _benefit_detail returns {measure_index: "'Benefit Detail'!B<row>"}.
    has_formula_components = (
        any(m.get("benefit_components") for m in data.get("measures", []))
        and data.get("specialist_type") not in ("natural_shading", "green_roof")
    )
    bd_total_refs = {}
    if has_formula_components:
        bd_total_refs = _benefit_detail(wb, data, rm=None)

    rm = _inputs(wb, data, am, bd_total_refs=bd_total_refs)
    _results(wb, data, rm)
    _sensitivity(wb, data, rm)
    _calculations(wb, data, rm, am)
    _summary(wb, data, rm)

    # Legacy benefit_types mode: build display-only Benefit Detail after Inputs
    if not has_formula_components and any(
        m.get("benefit_types") for m in data.get("measures", [])
    ):
        _benefit_detail(wb, data, rm)

    if data.get("specialist_type") in ("natural_shading", "green_roof"):
        _specialist_detail(wb, data, rm, am)
        _benefit_breakdown(wb, data, rm)
    wb.save(path)


# ── SHEET 1: INPUTS ────────────────────────────────────────────────────────────
def _inputs(wb, data, am=None, bd_total_refs=None):
    ws = wb.active
    ws.title = "Inputs"
    ws.sheet_view.showGridLines = False

    measures = data["measures"]
    n = len(measures)
    cur = f"{data['currency']} ({data['currency_unit']})"

    _hdr(ws, 1, 1, f"CLIMATE ADAPTATION CBA — {data['problem_title'].upper()}", sz=13, span=n+4)
    _hdr(ws, 2, 1, data["problem_summary"], bg=C_MID, fg="CBD5E1", bold=False, sz=9, wrap=True, span=n+4)
    ws.row_dimensions[2].height = 36
    _add_legend(ws, 3, 1, span=4)

    row = 10
    _sec(ws, row, 1, "GLOBAL PARAMETERS", span=4); row += 1

    _cell(ws, row, 1, "Discount Rate", bold=True)
    _glob_cell(ws, row, 2, data["discount_rate"], fmt="0.00%")
    _cell(ws, row, 3, "← change here to update all calculations", color="94A3B8")
    DR_ROW = row; row += 1

    _cell(ws, row, 1, "Time Horizon (years)", bold=True)
    _glob_cell(ws, row, 2, data["time_horizon"], fmt="#,##0")
    _cell(ws, row, 3, "← change here to update all calculations", color="94A3B8")
    YR_ROW = row; row += 1

    _cell(ws, row, 1, "Currency", bold=True)
    _cell(ws, row, 2, cur)
    row += 2

    _sec(ws, row, 1, f"MEASURE INPUTS  [{cur}]  —  🔵 Blue = local input  |  🟠 Amber = literature value", span=n+3); row += 1

    _hdr(ws, row, 1, "Parameter", bg=C_ACCENT, sz=10)
    for ci, m in enumerate(measures, 2):
        _hdr(ws, row, ci, m["name"], bg=C_ACCENT, sz=10)
    _hdr(ws, row, n+2, "Notes", bg=C_ACCENT, sz=10)
    row += 1

    CAPEX_ROW = row
    _cell(ws, row, 1, "Capital Cost / CAPEX", bold=True)
    for ci, m in enumerate(measures, 2):
        _endo_cell(ws, row, ci, m["capex"], fmt="#,##0.0")
    _cell(ws, row, n+2, "One-time capital cost — local estimate"); row += 1

    OPEX_ROW = row
    _cell(ws, row, 1, "Annual O&M Cost / OPEX", bold=True)
    for ci, m in enumerate(measures, 2):
        _endo_cell(ws, row, ci, m["annual_opex"], fmt="#,##0.0")
    _cell(ws, row, n+2, "Recurring annual cost — local estimate"); row += 1

    BENEFIT_ROW = row
    _cell(ws, row, 1, "Annual Benefit", bold=True)
    sub_total_ref = am.get("sub_total") if (am and data.get("specialist_type") in ("natural_shading", "green_roof")) else None
    if sub_total_ref:
        # Specialist mode: formula-linked from ASSUMPTIONS total — computed green
        for ci, m in enumerate(measures, 2):
            c = ws.cell(row=row, column=ci, value=f"={sub_total_ref}")
            c.font = Font(name="Arial", bold=True, color=GREEN_LK, size=10)
            c.number_format = "#,##0.0"; c.border = _bd()
            c.fill = PatternFill("solid", fgColor=C_COMP_BG)
            c.alignment = Alignment(horizontal="right")
        _cell(ws, row, n+2, "← derived from ASSUMPTIONS benefit components (see CALCULATIONS sheet for breakdown)")
    else:
        # Generic / formula-engine mode
        bd_refs = bd_total_refs or {}
        has_formulas = bool(bd_refs)
        for idx, (ci, m) in enumerate(zip(range(2, n+2), measures)):
            ref = bd_refs.get(idx)
            if ref:
                # Link live to Benefit Detail total — the KEY fix
                c = ws.cell(row=row, column=ci, value=f"={ref}")
                c.font   = Font(name="Arial", bold=True, color=GREEN_LK, size=10)
                c.fill   = PatternFill("solid", fgColor=C_COMP_BG)
                c.number_format = "#,##0.000"
                c.border = _bd()
                c.alignment = Alignment(horizontal="right")
            else:
                # Fallback: static value from JSON
                benefit_val = m.get("annual_benefit") or 0
                _exog_cell(ws, row, ci, benefit_val, fmt="#,##0.0")
        note = ("🟢 Computed in Benefit Detail sheet — see formula breakdown per component"
                if has_formulas else
                "🟠 Exogenous — literature-sourced estimate (editable)")
        _cell(ws, row, n+2, note)
    row += 1

    # Benefit component expansion — specialist only, inserted between Annual Benefit and Lifetime
    if am and am.get("sub_total") and data.get("specialist_type") in ("natural_shading", "green_roof"):
        COMP_LABELS = [
            ("sub_avoided_mortality",      "  \u2937 Avoided Mortality",            "NIS M/yr  |  step-by-step in ASSUMPTIONS"),
            ("sub_morbidity_savings",      "  \u2937 + Morbidity Savings",          "NIS M/yr  |  step-by-step in ASSUMPTIONS"),
            ("sub_skin_cancer_prevention", "  \u2937 + Skin Cancer Prevention",     "NIS M/yr  |  natural shading"),
            ("sub_carbon_sequestration",   "  \u2937 + Carbon Sequestration",       "NIS M/yr"),
            ("sub_runoff_reduction",       "  \u2937 + Runoff Reduction",           "NIS M/yr"),
            ("sub_air_quality",            "  \u2937 + Air Quality",                "NIS M/yr"),
            ("sub_habitat_creation",       "  \u2937 + Habitat Creation",           "NIS M/yr"),
            ("sub_property_value_uplift",  "  \u2937 + Property Value Uplift",      "Year-1 lump sum  |  green roof"),
            ("sub_roof_longevity",         "  \u2937 + Roof Longevity Extension",   "Year-1 lump sum  |  green roof"),
        ]
        # Section header bar
        c = ws.cell(row=row, column=1,
                    value="ANNUAL BENEFIT COMPONENTS  —  each value is an Excel formula linked from ASSUMPTIONS sheet")
        c.font = Font(name="Arial", bold=True, color=C_WHITE, size=8)
        c.fill = PatternFill("solid", fgColor=C_MID)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = _bd()
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n+2)
        ws.row_dimensions[row].height = 14
        row += 1
        # Component rows
        for key, label, btype in COMP_LABELS:
            ref = am.get(key)
            if not ref:
                continue
            c1 = ws.cell(row=row, column=1, value=label)
            c1.font = Font(name="Arial", size=8, color="64748B", italic=True)
            c1.border = _bd()
            for ci_m in range(2, n+2):
                c2 = ws.cell(row=row, column=ci_m, value=f"={ref}")
                c2.font = Font(name="Arial", size=8, color=GREEN_LK)
                c2.number_format = "#,##0.000"; c2.border = _bd()
                c2.alignment = Alignment(horizontal="right")
            c3 = ws.cell(row=row, column=n+2, value=btype)
            c3.font = Font(name="Arial", size=8, color="94A3B8", italic=True)
            c3.border = _bd()
            ws.row_dimensions[row].height = 13
            row += 1
        # Total row
        c1 = ws.cell(row=row, column=1, value="  \u25ba TOTAL ANNUAL BENEFIT")
        c1.font = Font(name="Arial", bold=True, size=9, color=C_GREEN)
        c1.fill = PatternFill("solid", fgColor="F0FDF4"); c1.border = _bd()
        for ci_m in range(2, n+2):
            c2 = ws.cell(row=row, column=ci_m, value=f"={am['sub_total']}")
            c2.font = Font(name="Arial", bold=True, size=9, color=C_GREEN)
            c2.number_format = "#,##0.000"
            c2.fill = PatternFill("solid", fgColor="F0FDF4"); c2.border = _bd()
            c2.alignment = Alignment(horizontal="right")
        c3 = ws.cell(row=row, column=n+2, value="= Annual Benefit row above")
        c3.font = Font(name="Arial", size=8, color="94A3B8", italic=True)
        c3.fill = PatternFill("solid", fgColor="F0FDF4"); c3.border = _bd()
        ws.row_dimensions[row].height = 14
        row += 2  # blank gap before Lifetime row

    LIFE_ROW = row
    _cell(ws, row, 1, "Measure Lifetime (years)", bold=True)
    for ci, m in enumerate(measures, 2):
        c = ws.cell(row=row, column=ci, value=m["lifetime_years"])
        c.font = Font(name="Arial", bold=True, color=BLUE, size=10)
        c.number_format = "#,##0"; c.border = _bd()
        c.fill = PatternFill("solid", fgColor=C_LIGHT)
        c.alignment = Alignment(horizontal="right")
    _cell(ws, row, n+2, "Used for PV calculations"); row += 1

    for label, key in [("Category", "category"), ("Feasibility", "feasibility"),
                       ("Uncertainty", "uncertainty"), ("Data Source", "data_source")]:
        _cell(ws, row, 1, label, bold=True)
        for ci, m in enumerate(measures, 2):
            _cell(ws, row, ci, m.get(key, ""), align="center")
        row += 1

    _cell(ws, row, 1, "Benefit Types", bold=True)
    for ci, m in enumerate(measures, 2):
        btypes = m.get("benefit_types") or [
            bc.get("name","") for bc in (m.get("benefit_components") or [])
        ]
        _cell(ws, row, ci, ", ".join(btypes), align="center")
    row += 1

    _cell(ws, row, 1, "Co-benefits", bold=True)
    for ci, m in enumerate(measures, 2):
        _cell(ws, row, ci, m["co_benefits"])
    row += 2

    _sec(ws, row, 1, "KEY ASSUMPTIONS & LIMITATIONS", span=n+3); row += 1
    # Header row for assumptions table
    _hdr(ws, row, 1, "#", bg=C_ACCENT, sz=9)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=n+2)
    _hdr(ws, row, 2, "Assumption / Limitation", bg=C_ACCENT, sz=9)
    _hdr(ws, row, n+3, "Source", bg=C_ACCENT, sz=9)
    row += 1
    # Parse assumptions — list of {text, source} or plain string
    raw_assumptions = data.get("key_assumptions", "")
    if isinstance(raw_assumptions, list):
        assumption_list = raw_assumptions
    else:
        # Split plain string into individual assumptions
        import re
        parts = re.split(r'(?<=[.;])\s+', str(raw_assumptions))
        assumption_list = [{"text": p.strip(" ;."), "source": ""} for p in parts if p.strip(" ;.")]
    for idx, item in enumerate(assumption_list, 1):
        text   = item.get("text", item) if isinstance(item, dict) else str(item)
        source = item.get("source", "") if isinstance(item, dict) else ""
        _cell(ws, row, 1, str(idx), bold=True, align="center")
        c = ws.cell(row=row, column=2, value=text)
        c.font = Font(name="Arial", size=9, color="475569")
        c.alignment = Alignment(wrap_text=True, vertical="top")
        c.border = _bd()
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=n+2)
        c2 = ws.cell(row=row, column=n+3, value=source)
        c2.font = Font(name="Arial", size=9, color="475569", italic=True)
        c2.alignment = Alignment(wrap_text=True, vertical="top")
        c2.border = _bd()
        ws.row_dimensions[row].height = 28
        row += 1
    row += 1

    _cell(ws, row, 1, "Data Gaps:", bold=True)
    c = ws.cell(row=row, column=2, value=data.get("data_gaps", ""))
    c.font = Font(name="Arial", size=9, color="475569")
    c.alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=n+3); row += 2

    _add_legend(ws, row, 1, span=4); row += 8

    _auto_widths(ws)

    return {"dr": DR_ROW, "yr": YR_ROW,
            "capex": CAPEX_ROW, "opex": OPEX_ROW,
            "benefit": BENEFIT_ROW, "life": LIFE_ROW, "n": n}


# ── SHEET 2: CBA RESULTS ───────────────────────────────────────────────────────
def _results(wb, data, rm):
    ws = wb.create_sheet("CBA Results")
    ws.sheet_view.showGridLines = False
    measures = data["measures"]
    n = rm["n"]

    _hdr(ws, 1, 1, "COST-BENEFIT ANALYSIS — RESULTS", sz=13, span=n+3)
    _hdr(ws, 2, 1,
         f"All values in {data['currency']} {data['currency_unit']}  |  Change blue cells in Inputs sheet to update",
         bg=C_MID, fg="CBD5E1", bold=False, sz=9, span=n+3)
    _add_legend(ws, 3, 1, span=4)

    row = 10
    _hdr(ws, row, 1, "Metric", bg=C_ACCENT, sz=10)
    for ci, m in enumerate(measures, 2):
        _hdr(ws, row, ci, m["name"], bg=C_ACCENT, sz=10)
    _hdr(ws, row, n+2, "Best", bg=C_ACCENT, sz=10)
    row += 1

    dr   = f"Inputs!$B${rm['dr']}"
    # Helper references
    def inp(key, ci):
        return f"Inputs!{get_column_letter(ci)}${rm[key]}"

    rr = {}  # track row numbers

    def add_row(label, key, f_fn, fmt, is_key=False):
        nonlocal row
        _cell(ws, row, 1, label, bold=is_key)
        for ci in range(2, n+2):
            f = f_fn(ci)
            c = ws.cell(row=row, column=ci, value=f)
            is_inp_link = f.startswith("=Inputs")
            if is_key:
                # Key computed results (NPV, BCR) — green background
                c.font = Font(name="Arial", color=C_COMP_FG, bold=True, size=10)
                c.fill = PatternFill("solid", fgColor=C_COMP_BG)
            elif is_inp_link:
                # Cross-sheet input references — green text
                c.font = Font(name="Arial", color=GREEN_LK, bold=False, size=10)
                c.fill = PatternFill("solid", fgColor="E8F5E9")
            else:
                # Intermediate computed — light grey
                c.font = Font(name="Arial", color="475569", bold=False, size=10)
                c.fill = PatternFill("solid", fgColor="F8FAFC")
            c.number_format = fmt; c.border = _bd()
            c.alignment = Alignment(horizontal="right", vertical="center")
        rr[key] = row; row += 1

    add_row(f"CAPEX ({data['currency']}M)", "capex",
            lambda ci: f"=Inputs!{get_column_letter(ci)}${rm['capex']}", "#,##0.0")
    add_row(f"Annual Benefit ({data['currency']}M)", "ann_ben",
            lambda ci: f"=Inputs!{get_column_letter(ci)}${rm['benefit']}", "#,##0.0")
    add_row(f"Annual OPEX ({data['currency']}M)", "ann_opex",
            lambda ci: f"=Inputs!{get_column_letter(ci)}${rm['opex']}", "#,##0.0")

    def pv_ben_f(ci):
        b = inp("benefit", ci); l = inp("life", ci)
        return f"={b}/{dr}*(1-(1+{dr})^(-{l}))"

    def pv_cost_f(ci):
        k = inp("capex", ci); o = inp("opex", ci); l = inp("life", ci)
        return f"={k}+{o}/{dr}*(1-(1+{dr})^(-{l}))"

    add_row(f"PV of Benefits ({data['currency']}M)", "pv_ben", pv_ben_f, "#,##0.0")
    PVB = rr["pv_ben"]
    add_row(f"PV of Costs ({data['currency']}M)", "pv_cost", pv_cost_f, "#,##0.0")
    PVC = rr["pv_cost"]

    def npv_f(ci):
        col = get_column_letter(ci)
        return f"={col}{PVB}-{col}{PVC}"

    add_row(f"NPV ({data['currency']}M)", "npv", npv_f,
            '#,##0.0;(#,##0.0);"-"', is_key=True)
    NPV_ROW = rr["npv"]

    def bcr_f(ci):
        col = get_column_letter(ci)
        return f"=IF({col}{PVC}>0,{col}{PVB}/{col}{PVC},0)"

    add_row("BCR", "bcr", bcr_f, "0.00", is_key=True)
    BCR_ROW = rr["bcr"]

    def payback_f(ci):
        k = inp("capex", ci); b = inp("benefit", ci); o = inp("opex", ci)
        return f'=IF(({b}-{o})>0,{k}/({b}-{o}),"N/A")'

    add_row("Simple Payback (years)", "payback", payback_f, "0.0")

    def ann_net_f(ci):
        b = inp("benefit", ci); o = inp("opex", ci)
        return f"={b}-{o}"

    add_row(f"Annual Net Benefit ({data['currency']}M)", "ann_net", ann_net_f, "#,##0.0")

    # Best column formulas
    bcr_range = f"{get_column_letter(2)}{BCR_ROW}:{get_column_letter(n+1)}{BCR_ROW}"
    npv_range = f"{get_column_letter(2)}{NPV_ROW}:{get_column_letter(n+1)}{NPV_ROW}"

    for key_r, rng, fmt in [(BCR_ROW, bcr_range, "0.00"), (NPV_ROW, npv_range, "#,##0.0")]:
        c = ws.cell(row=key_r, column=n+2, value=f"=MAX({rng})")
        c.font = Font(name="Arial", color=C_GREEN, bold=True, size=10)
        c.number_format = fmt; c.border = _bd()
        c.fill = PatternFill("solid", fgColor=C_LIGHT)
        c.alignment = Alignment(horizontal="right")

    # Best measure name (on the header row)
    HDR_ROW = 10
    name_range = f"{get_column_letter(2)}{HDR_ROW}:{get_column_letter(n+1)}{HDR_ROW}"
    c = ws.cell(row=HDR_ROW, column=n+2,
                value=f'=INDEX({name_range},MATCH(MAX({bcr_range}),{bcr_range},0))')
    c.font = Font(name="Arial", color=C_GREEN, bold=True, size=10)
    c.fill = PatternFill("solid", fgColor=C_DARK)
    c.alignment = Alignment(horizontal="center")

    # Ranking & Viability
    row += 1
    _sec(ws, row, 1, "RANKING & VIABILITY", span=n+2); row += 1

    _cell(ws, row, 1, "BCR Rank (1 = best)", bold=True)
    for ci in range(2, n+2):
        col = get_column_letter(ci)
        c = ws.cell(row=row, column=ci,
                    value=f"=RANK({col}{BCR_ROW},{bcr_range},0)")
        c.font = Font(name="Arial", color=BLACK, bold=True, size=10)
        c.number_format = "0"; c.border = _bd()
        c.alignment = Alignment(horizontal="center")
    row += 1

    _cell(ws, row, 1, "Viable? (BCR ≥ 1)", bold=True)
    for ci in range(2, n+2):
        col = get_column_letter(ci)
        c = ws.cell(row=row, column=ci,
                    value=f'=IF({col}{BCR_ROW}>=1,"✓ Yes","✗ No")')
        c.font = Font(name="Arial", color=BLACK, size=10)
        c.border = _bd()
        c.alignment = Alignment(horizontal="center")
    row += 2

    # Chart data area
    chart_r = row
    ws.cell(row=chart_r, column=1, value="Measure")
    ws.cell(row=chart_r, column=2, value="BCR")
    for mi, m in enumerate(measures, 1):
        ws.cell(row=chart_r+mi, column=1, value=m["name"])
        ws.cell(row=chart_r+mi, column=2,
                value=f"={get_column_letter(mi+1)}{BCR_ROW}")

    chart = BarChart()
    chart.type = "col"; chart.title = "BCR by Measure"
    chart.y_axis.title = "Benefit-Cost Ratio"
    chart.style = 10; chart.width = 18; chart.height = 12

    data_ref = Reference(ws, min_col=2, max_col=2, min_row=chart_r, max_row=chart_r+n)
    cats_ref = Reference(ws, min_col=1, min_row=chart_r+1, max_row=chart_r+n)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    ws.add_chart(chart, f"A{chart_r+n+3}")

    _auto_widths(ws)

    # Expose BCR_ROW for summary sheet
    rm["results_bcr_row"]   = BCR_ROW
    rm["results_capex_row"] = rr["capex"]
    rm["results_npv_row"]   = NPV_ROW


def _build_specialist_sensitivity_vars(data):
    """Return the 4-pillar sensitivity variable list for specialist measure types."""
    sp  = data.get("specialist_params", {})
    dr  = data.get("discount_rate", 0.035)
    eff = sp.get("heat_reduction_efficiency", 0.4)
    return [
        {"name": "Discount Rate",            "low": 0.02,             "base": dr,  "high": 0.10,            "unit": "%"},
        {"name": "Heat Reduction Efficiency","low": round(eff*0.7,3), "base": eff, "high": round(eff*1.3,3),"unit": "multiplier"},
        {"name": "CAPEX Variation",          "low": 0.80,             "base": 1.00,"high": 1.30,            "unit": "multiplier"},
        {"name": "Activity Level",           "low": 0.70,             "base": 1.00,"high": 1.30,            "unit": "multiplier"},
    ]


# ── SHEET 3: SENSITIVITY ──────────────────────────────────────────────────────
def _sensitivity(wb, data, rm):
    ws = wb.create_sheet("Sensitivity")
    ws.sheet_view.showGridLines = False
    measures = data["measures"]
    n = rm["n"]
    sens_vars = data.get("sensitivity_vars", [])
    if data.get("specialist_type") in ("natural_shading", "green_roof"):
        sens_vars = _build_specialist_sensitivity_vars(data)
    # Ensure generic mode always has meaningful sensitivity vars
    if not sens_vars:
        dr = data.get("discount_rate", 0.035)
        sens_vars = [
            {"name": "Discount Rate",             "low": 0.02,  "base": dr,   "high": 0.10, "unit": "%"},
            {"name": "Annual Benefit Multiplier", "low": 0.60,  "base": 1.00, "high": 1.40, "unit": "multiplier"},
            {"name": "CAPEX Variation",           "low": 0.80,  "base": 1.00, "high": 1.30, "unit": "multiplier"},
        ]

    _hdr(ws, 1, 1, "SENSITIVITY ANALYSIS", sz=13, span=n+5)
    _hdr(ws, 2, 1,
         "Each parameter is analysed independently. Edit the blue cells in each range table; BCR results update automatically.",
         bg=C_MID, fg="CBD5E1", bold=False, sz=9, span=n+5)

    dr_inp = f"Inputs!$B${rm['dr']}"
    row = 4

    for sv in sens_vars:
        sv_name = sv["name"]

        # ── Parameter section header ───────────────────────────────────────────
        _sec(ws, row, 1, f"PARAMETER: {sv_name.upper()}", span=n+4); row += 1

        # ── Range table (Variable | Low | Base | High | Unit) ─────────────────
        for ci, h in enumerate(["Variable", "Low", "Base", "High", "Unit"], 1):
            _hdr(ws, row, ci, h, bg=C_ACCENT, sz=10)
        row += 1

        _cell(ws, row, 1, sv_name, bold=True)
        for col_i, val in enumerate([sv["low"], sv["base"], sv["high"]], 2):
            c = ws.cell(row=row, column=col_i, value=val)
            c.font = Font(name="Arial", bold=True, color=BLUE, size=10)
            c.number_format = "0.00%" if sv["unit"] == "%" else "0.00"
            c.border = _bd()
            c.fill = PatternFill("solid", fgColor=C_LIGHT)
            c.alignment = Alignment(horizontal="right")
        _cell(ws, row, 5, sv["unit"])
        sv_r = {
            "low":  f"$B${row}",
            "base": f"$C${row}",
            "high": f"$D${row}"
        }
        # Expose parameter row number so CALCULATIONS sheet can reference these cells
        rm.setdefault("sens_param_rows", {})[sv_name] = int(sv_r["low"][3:])
        row += 2  # blank row between range table and BCR table

        # ── BCR sensitivity table for this parameter ───────────────────────────
        _hdr(ws, row, 1, "Scenario", bg=C_ACCENT, sz=10)
        for ci, m in enumerate(measures, 2):
            _hdr(ws, row, ci, m["name"], bg=C_ACCENT, sz=10)
        row += 1

        for scenario, col_key, color in [("Low",  "low",  C_RED),
                                          ("Base", "base", BLACK),
                                          ("High", "high", C_GREEN)]:
            param_ref = sv_r.get(col_key, "$C$7")
            bg = C_LIGHT if scenario == "Base" else None

            _cell(ws, row, 1, scenario, color=color, bold=(scenario == "Base"), bg=bg)

            for mi in range(n):
                ci = mi + 2
                capex = f"Inputs!{get_column_letter(ci)}${rm['capex']}"
                opex  = f"Inputs!{get_column_letter(ci)}${rm['opex']}"
                ben   = f"Inputs!{get_column_letter(ci)}${rm['benefit']}"
                life  = f"Inputs!{get_column_letter(ci)}${rm['life']}"

                pv_cost_override = None
                if sv_name == "Discount Rate":
                    dr_use   = param_ref
                    ben_term = ben
                elif sv_name in ("Annual Benefit", "Annual Benefit Multiplier"):
                    dr_use   = dr_inp
                    ben_term = f"({ben}*{param_ref})"
                elif sv_name == "CAPEX Variation":
                    dr_use   = dr_inp
                    ben_term = ben
                    pv_cost_override = f"(({capex}*{param_ref})+{opex}/{dr_inp}*(1-(1+{dr_inp})^(-{life})))"
                elif sv_name == "Activity Level":
                    dr_use   = dr_inp
                    ben_term = f"({ben}*{param_ref})"
                elif sv_name == "Heat Reduction Efficiency":
                    dr_use   = dr_inp
                    base_eff = data.get("specialist_params", {}).get("heat_reduction_efficiency", 0.5)
                    ben_term = f"({ben}*{param_ref}/{base_eff})"
                else:
                    dr_use   = dr_inp
                    ben_term = ben

                pv_ben  = f"({ben_term}/{dr_use}*(1-(1+{dr_use})^(-{life})))"
                pv_cost = pv_cost_override if pv_cost_override else \
                          f"({capex}+{opex}/{dr_use}*(1-(1+{dr_use})^(-{life})))"
                formula = f"=IF({dr_use}>0,IF({pv_cost}>0,{pv_ben}/{pv_cost},0),0)"

                c = ws.cell(row=row, column=mi+2, value=formula)
                c.font = Font(name="Arial", color=color, bold=(scenario == "Base"), size=10)
                c.number_format = "0.00"; c.border = _bd()
                c.alignment = Alignment(horizontal="right")
                if bg: c.fill = PatternFill("solid", fgColor=C_LIGHT)
            row += 1

        row += 3  # blank rows before next parameter

    _sec(ws, row, 1, "INSTRUCTIONS", span=6); row += 1
    for inst in [
        "1. Each parameter block above has its own range table (blue editable cells) followed by BCR results",
        "2. BCR values recalculate automatically — no manual action needed",
        "3. 'Discount Rate': enter as decimal (e.g. 0.02 = 2%)",
        "4. Multiplier parameters: 1.0 = base case, 0.5 = 50% of base, 1.5 = 150%",
        "5. Parameters are analysed one at a time (all-else-equal); results are not combined",
    ]:
        c = ws.cell(row=row, column=1, value=inst)
        c.font = Font(name="Arial", size=9, color="475569")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        row += 1

    _auto_widths(ws)


# ── SHEET 4: SUMMARY ──────────────────────────────────────────────────────────
def _summary(wb, data, rm):
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False
    measures = data["measures"]
    n = rm["n"]
    BCR_ROW   = rm["results_bcr_row"]
    CAPEX_ROW = rm["results_capex_row"]
    NPV_ROW   = rm["results_npv_row"]

    _hdr(ws, 1, 1, "EXECUTIVE SUMMARY", sz=13, span=5)
    _hdr(ws, 2, 1, data["problem_title"], bg=C_MID, fg="CBD5E1", sz=11, span=5)

    row = 4
    _sec(ws, row, 1, "PROBLEM OVERVIEW", span=5); row += 1
    c = ws.cell(row=row, column=1, value=data["problem_summary"])
    c.font = Font(name="Arial", size=10)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=row, start_column=1, end_row=row+1, end_column=5)
    ws.row_dimensions[row].height = 40; row += 3

    _sec(ws, row, 1, "MEASURES AT A GLANCE  (values link live from CBA Results)", span=5); row += 1
    for ci, h in enumerate(["Measure", "Category", f"CAPEX ({data['currency']}M)",
                             "BCR (base)", "Viability"], 1):
        _hdr(ws, row, ci, h, bg=C_ACCENT, sz=10)
    row += 1

    for mi, m in enumerate(measures):
        ci = mi + 2  # Excel column in Results sheet
        col = get_column_letter(ci)

        _cell(ws, row, 1, m["name"], bold=True)
        _cell(ws, row, 2, m["category"], align="center")

        c = ws.cell(row=row, column=3, value=f"='CBA Results'!{col}{CAPEX_ROW}")
        c.font = Font(name="Arial", color=GREEN_LK, size=10)
        c.number_format = "#,##0.0"; c.border = _bd()
        c.alignment = Alignment(horizontal="right")

        c = ws.cell(row=row, column=4, value=f"='CBA Results'!{col}{BCR_ROW}")
        c.font = Font(name="Arial", color=GREEN_LK, bold=True, size=10)
        c.number_format = "0.00"; c.border = _bd()
        c.fill = PatternFill("solid", fgColor=C_LIGHT)
        c.alignment = Alignment(horizontal="center")

        bcr_here = f"D{row}"
        c = ws.cell(row=row, column=5,
                    value=f'=IF({bcr_here}>=1.5,"✓ Recommended",IF({bcr_here}>=1.0,"○ Consider","✗ Review"))')
        c.font = Font(name="Arial", color=BLACK, size=10)
        c.border = _bd(); c.alignment = Alignment(horizontal="center")
        row += 1

    row += 1
    _sec(ws, row, 1, "HOW TO USE THIS WORKBOOK", span=5); row += 1
    for step in [
        "1.  INPUTS sheet — Edit blue cells (CAPEX, OPEX, Benefits, Discount Rate, Time Horizon)",
        "2.  CBA RESULTS sheet — NPV and BCR recalculate automatically from Inputs",
        "3.  BENEFIT BREAKDOWN sheet — Shows NPV of each benefit category (mortality, morbidity, skin cancer, ecosystem services, property uplift, roof longevity) and their shares of total benefits",
        "4.  SENSITIVITY sheet — Edit Low/Base/High ranges; BCR table updates automatically",
        "5.  SPECIALIST DETAIL sheet — Inspect VSL derivation, CDD parameters, specialist parameters, and year-by-year benefit projections",
        "6.  SUMMARY sheet (this sheet) — BCR and CAPEX values link live from CBA Results; use together with Benefit Breakdown and Specialist Detail when presenting results",
        "7.  Never overwrite black or green cells — they contain formulas",
    ]:
        c = ws.cell(row=row, column=1, value=step)
        c.font = Font(name="Arial", size=10)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        row += 1

    _auto_widths(ws)


# ── SHEET 5: SPECIALIST DETAIL ────────────────────────────────────────────────
def _specialist_detail(wb, data, rm, am=None):
    if am is None:
        am = {}
    ws = wb.create_sheet("Specialist Detail")
    ws.sheet_view.showGridLines = False

    stype    = data.get("specialist_type", "")
    vsl      = data.get("vsl_params", {})
    cdd      = data.get("cdd_params", {})
    sp       = data.get("specialist_params", {})
    measures = data["measures"]
    n        = rm["n"]
    cur      = data.get("currency", "NIS")

    stype_label = "Natural Shading (Boulevard Trees)" if stype == "natural_shading" else "Green Roof"

    # ── Row 1-2: Header ────────────────────────────────────────────────────────
    _hdr(ws, 1, 1, f"SPECIALIST METHODOLOGY DETAIL — {stype_label.upper()}", sz=13, span=10)
    _hdr(ws, 2, 1,
         f"Functional unit: {sp.get('functional_unit','—')}  |  {data.get('problem_title','')}",
         bg=C_MID, fg="CBD5E1", bold=False, sz=9, span=10)

    row = 4

    # ── Section: VSL Derivation ────────────────────────────────────────────────
    _sec(ws, row, 1, "VALUE OF STATISTICAL LIFE — STEP-BY-STEP DERIVATION", span=4); row += 1
    for ci, h in enumerate(["Step", "Parameter", "Value", "Notes"], 1):
        _hdr(ws, row, ci, h, bg=C_ACCENT, sz=10)
    row += 1

    # Build VSL rows using ASSUMPTIONS formula references where available,
    # falling back to Python-computed static values if am is empty.
    def _vsl_val(am_key, fallback):
        ref = am.get(am_key)
        return f"={ref}" if ref else fallback

    VSL_ROWS = [
        ("1",  "Base VSL (OECD, 2005 USD)",        _vsl_val("vsl_base",    vsl.get("base_vsl_usd_2005", 3_000_000)), "#,##0",  "OECD meta-study baseline"),
        ("2",  f"× CPI Multiplier (2005→2024)",     _vsl_val("cpi_mult",    vsl.get("cpi_multiplier", 1.68)),         "0.000",  "US Bureau of Labor Statistics"),
        ("3",  "= CPI-Adjusted VSL (2024 USD)",     _vsl_val("vsl_cpi_adj", round(vsl.get("base_vsl_usd_2005", 3_000_000) * vsl.get("cpi_multiplier", 1.68))), "#,##0", ""),
        ("4",  f"× GDP PPP Ratio (Israel / OECD)",  _vsl_val("ppp_ratio",   vsl.get("gdp_ppp_ratio", 0.89)),          "0.000",  "World Bank WDI"),
        ("5",  "× Income Elasticity",               _vsl_val("income_el",   vsl.get("income_elasticity", 1.0)),        "0.0",    "Standard for developed economies"),
        ("6",  "= PPP-Adjusted VSL (USD)",          _vsl_val("vsl_ppp_adj", round(vsl.get("base_vsl_usd_2005", 3_000_000) * vsl.get("cpi_multiplier", 1.68) * vsl.get("gdp_ppp_ratio", 0.89) * vsl.get("income_elasticity", 1.0))), "#,##0", ""),
        ("7",  f"× Exchange Rate ({cur}/USD)",       _vsl_val("fx_rate",     vsl.get("usd_to_local_currency", 3.7)),   "0.00",   ""),
        ("8",  f"= VSL in {cur}",                   _vsl_val("vsl_local",   vsl.get("computed_vsl_local", 12_800_000)), "#,##0", "Key output — used in benefit formulas"),
        ("9",  "÷ Remaining Life Expectancy (yrs)", _vsl_val("life_exp",    vsl.get("life_expectancy_remaining", 35)), "0",      "Affected demographic"),
        ("10", f"= VSLY in {cur}",                  _vsl_val("vsly_local",  vsl.get("computed_vsly_local", 365_714)),  "#,##0",  "Value per Statistical Life Year"),
    ]
    # Rows that are computed (show in green); others are input references (show in green-link)
    COMPUTED_STEPS = {"3", "6", "8", "10"}
    for step, param, val, fmt, note in VSL_ROWS:
        is_computed = step in COMPUTED_STEPS
        _cell(ws, row, 1, step, align="center")
        _cell(ws, row, 2, param, bold=is_computed)
        c = ws.cell(row=row, column=3, value=val)
        c.font = Font(name="Arial", bold=is_computed,
                      color=C_GREEN if is_computed else GREEN_LK, size=10)
        c.number_format = fmt; c.border = _bd()
        c.alignment = Alignment(horizontal="right")
        if is_computed:
            c.fill = PatternFill("solid", fgColor="F0FDF4")
        _cell(ws, row, 4, note, color="94A3B8")
        row += 1

    row += 1

    # ── Section: CDD Parameters ────────────────────────────────────────────────
    _sec(ws, row, 1, "COOLING DEGREE DAY PARAMETERS", span=4); row += 1
    for ci, h in enumerate(["Parameter", "Value", "Unit", "Notes"], 1):
        _hdr(ws, row, ci, h, bg=C_ACCENT, sz=10)
    row += 1

    pop_label = "Pedestrians per Hour" if stype == "natural_shading" else "Population Density (people/km²)"
    CDD_ROWS = [
        ("Annual Cooling Degree Days",   cdd.get("annual_cdd", 735),          "CDD",        "Tel Aviv baseline (21°C base)"),
        ("Base Temperature",             cdd.get("base_temp_celsius", 21),     "°C",         "Threshold for heat health risk"),
        ("Heat-Mortality Factor",        cdd.get("heat_mortality_factor", 0.00083), "deaths/person", "Per CDD above threshold"),
        (pop_label,                      cdd.get("population_density_or_pedestrians", "—"), "people", "Population at risk driver"),
    ]
    for param, val, unit, note in CDD_ROWS:
        _cell(ws, row, 1, param, bold=True)
        c = ws.cell(row=row, column=2, value=val)
        c.font = Font(name="Arial", color=BLACK, size=10)
        c.number_format = "0.00000" if isinstance(val, float) and val < 0.01 else "#,##0.##"
        c.border = _bd(); c.alignment = Alignment(horizontal="right")
        _cell(ws, row, 3, unit, align="center")
        _cell(ws, row, 4, note, color="94A3B8")
        row += 1

    row += 1

    # ── Section: Specialist Parameters ────────────────────────────────────────
    _sec(ws, row, 1, "SPECIALIST PARAMETERS", span=4); row += 1
    for ci, h in enumerate(["Parameter", "Value", "Unit"], 1):
        _hdr(ws, row, ci, h, bg=C_ACCENT, sz=10)
    row += 1

    if stype == "natural_shading":
        SP_ROWS = [
            ("Heat Reduction Efficiency", sp.get("heat_reduction_efficiency", 0.50), "ratio", True),
            ("UV Reduction Factor",       sp.get("uv_reduction_factor", 0.75),      "ratio", False),
            ("Pedestrians per Hour",      sp.get("pedestrians_per_hour", 1200),      "pax/hr", False),
            ("Functional Unit",           sp.get("functional_unit", "linear meter"), "—",    False),
        ]
    else:
        SP_ROWS = [
            ("Heat Reduction Efficiency", sp.get("heat_reduction_efficiency", 0.28), "ratio",  True),
            ("Property Value Uplift %",   sp.get("property_value_uplift_pct", 0.03), "%",      False),
            ("Roof Longevity Extension",  sp.get("roof_longevity_extension_years", 15), "years", False),
            ("Roof Area",                 sp.get("roof_area_m2", 1000),               "m²",    False),
            ("Functional Unit",           sp.get("functional_unit", "sq meter"),      "—",     False),
        ]

    for param, val, unit, is_editable in SP_ROWS:
        _cell(ws, row, 1, param, bold=True)
        c = ws.cell(row=row, column=2, value=val)
        if is_editable:
            c.font = Font(name="Arial", bold=True, color=BLUE, size=10)
            c.fill = PatternFill("solid", fgColor=C_LIGHT)
        else:
            c.font = Font(name="Arial", color=BLACK, size=10)
        c.number_format = "0.00%" if unit in ("%", "ratio") and isinstance(val, float) and val < 1 else "#,##0.##"
        c.border = _bd(); c.alignment = Alignment(horizontal="right")
        _cell(ws, row, 3, unit, align="center")
        row += 1

    row += 1

    # ── Section: Formula Drivers ───────────────────────────────────────────────
    _sec(ws, row, 1, "YEAR-BY-YEAR FORMULA DRIVERS  (blue = editable, green = linked from Inputs)", span=4); row += 1

    if stype == "natural_shading":
        mat_val = sp.get("maturity_years", 8)
        _cell(ws, row, 1, "Maturity Period (years)", bold=True)
        c = ws.cell(row=row, column=2, value=mat_val)
        c.font = Font(name="Arial", bold=True, color=BLUE, size=10)
        c.number_format = "0"; c.border = _bd()
        c.fill = PatternFill("solid", fgColor=C_LIGHT)
        c.alignment = Alignment(horizontal="right")
        _cell(ws, row, 3, "years")
        _cell(ws, row, 4, "← edit to change maturity ramp", color="94A3B8")
        MAT_ROW = row; row += 1
    else:
        MAT_ROW = None

    _cell(ws, row, 1, "Discount Rate (from Inputs)", bold=True)
    dr_formula = f"=Inputs!$B${rm['dr']}"
    c = ws.cell(row=row, column=2, value=dr_formula)
    c.font = Font(name="Arial", bold=True, color=GREEN_LK, size=10)
    c.number_format = "0.00%"; c.border = _bd()
    c.alignment = Alignment(horizontal="right")
    _cell(ws, row, 3, "%")
    _cell(ws, row, 4, "← change in Inputs sheet", color="94A3B8")
    DR_CELL_ROW = row; row += 2

    # ── Section: Year-by-Year Benefit Tables (one per measure) ────────────────
    _sec(ws, row, 1, "YEAR-BY-YEAR BENEFIT PROJECTION  (formulas — updates automatically)", span=10); row += 1

    # Benefit type definitions per specialist type
    if stype == "natural_shading":
        BEN_TYPES = [
            ("avoided_mortality_npv",      "Avoided Mortality"),
            ("morbidity_savings_npv",      "Morbidity"),
            ("skin_cancer_prevention_npv", "Skin Cancer Prev."),
            ("carbon_sequestration_npv",   "Carbon Seq."),
            ("runoff_reduction_npv",       "Runoff Reduction"),
            ("air_quality_npv",            "Air Quality"),
            ("habitat_creation_npv",       "Habitat"),
        ]
    else:
        BEN_TYPES = [
            ("avoided_mortality_npv",      "Avoided Mortality"),
            ("morbidity_savings_npv",      "Morbidity"),
            ("property_value_uplift_npv",  "Prop. Value Uplift"),
            ("roof_longevity_npv",        "Roof Longevity"),
            ("carbon_sequestration_npv",   "Carbon Seq."),
            ("runoff_reduction_npv",       "Runoff Reduction"),
            ("air_quality_npv",            "Air Quality"),
            ("habitat_creation_npv",       "Habitat"),
        ]

    n_ben = len(BEN_TYPES)

    # Map each benefit key to its ASSUMPTIONS annual sub-benefit reference
    # Keys use the full _npv suffix as in BEN_TYPES
    BEN_TYPE_AM = {
        "avoided_mortality_npv":      am.get("sub_avoided_mortality_npv",      am.get("sub_avoided_mortality",      "0")),
        "morbidity_savings_npv":      am.get("sub_morbidity_savings_npv",      am.get("sub_morbidity_savings",      "0")),
        "skin_cancer_prevention_npv": am.get("sub_skin_cancer_prevention_npv", am.get("sub_skin_cancer_prevention", "0")),
        "carbon_sequestration_npv":   am.get("sub_carbon_sequestration_npv",   am.get("sub_carbon_sequestration",   "0")),
        "runoff_reduction_npv":       am.get("sub_runoff_reduction_npv",       am.get("sub_runoff_reduction",       "0")),
        "air_quality_npv":            am.get("sub_air_quality_npv",            am.get("sub_air_quality",            "0")),
        "habitat_creation_npv":       am.get("sub_habitat_creation_npv",       am.get("sub_habitat_creation",       "0")),
        "property_value_uplift_npv":  am.get("sub_property_value_uplift_npv",  am.get("sub_property_value_uplift",  "0")),
        "roof_longevity_npv":         am.get("sub_roof_longevity_npv",         am.get("sub_roof_longevity",         "0")),
    }

    for mi, m in enumerate(measures):
        inp_col = get_column_letter(mi + 2)  # Inputs sheet column for this measure

        # Sub-table header
        _sec(ws, row, 1, f"Measure: {m['name']}", span=6 + n_ben); row += 1

        # Column headers (Year | Maturity | Base Benefit | Effective | Disc Factor | PV | benefit cols...)
        TABLE_HDRS = ["Year", "Maturity\nFactor", f"Base Benefit\n({cur}M/yr)",
                      f"Eff. Benefit\n({cur}M/yr)", "Disc.\nFactor", f"PV of Benefit\n({cur}M)"]
        for ci, h in enumerate(TABLE_HDRS, 1):
            _hdr(ws, row, ci, h, bg=C_ACCENT, sz=9, wrap=True)
        ws.row_dimensions[row].height = 30
        HDR_ROW = row; row += 1
        DATA_START = row

        mat_ref  = f"$B${MAT_ROW}" if MAT_ROW else None
        dr_ref   = f"$B${DR_CELL_ROW}"
        ben_ref  = f"Inputs!{inp_col}${rm['benefit']}"

        for yr in range(1, 51):
            r = row
            # Year
            c = ws.cell(row=r, column=1, value=yr)
            c.font = Font(name="Arial", size=9); c.border = _bd()
            c.alignment = Alignment(horizontal="center")

            # Maturity Factor
            if stype == "natural_shading" and mat_ref:
                mat_formula = f"=IF(A{r}<={mat_ref},A{r}/{mat_ref},1)"
            else:
                mat_formula = "=1"
            c = ws.cell(row=r, column=2, value=mat_formula)
            c.font = Font(name="Arial", color=BLACK, size=9); c.border = _bd()
            c.number_format = "0.00%"; c.alignment = Alignment(horizontal="right")

            # Base Annual Benefit
            c = ws.cell(row=r, column=3, value=f"={ben_ref}")
            c.font = Font(name="Arial", color=GREEN_LK, size=9); c.border = _bd()
            c.number_format = "#,##0.0"; c.alignment = Alignment(horizontal="right")

            # Effective Benefit
            c = ws.cell(row=r, column=4, value=f"=B{r}*C{r}")
            c.font = Font(name="Arial", color=BLACK, size=9); c.border = _bd()
            c.number_format = "#,##0.0"; c.alignment = Alignment(horizontal="right")

            # Discount Factor
            c = ws.cell(row=r, column=5, value=f"=1/(1+{dr_ref})^A{r}")
            c.font = Font(name="Arial", color=BLACK, size=9); c.border = _bd()
            c.number_format = "0.0000"; c.alignment = Alignment(horizontal="right")

            # PV of Benefit
            c = ws.cell(row=r, column=6, value=f"=D{r}*E{r}")
            c.font = Font(name="Arial", color=C_GREEN, bold=True, size=9); c.border = _bd()
            c.number_format = "#,##0.0"; c.alignment = Alignment(horizontal="right")
            c.fill = PatternFill("solid", fgColor="F0FDF4")

            # Benefit breakdown columns — reference ASSUMPTIONS sub-benefit annual values
            for bi, (key, _) in enumerate(BEN_TYPES):
                ann_ref = BEN_TYPE_AM.get(key, "0")
                if ann_ref == "0":
                    formula = "=0"
                elif key in ("property_value_uplift_npv", "roof_longevity_npv"):
                    # Lump-sum benefits: Year 1 only (discounted but no maturity factor)
                    formula = f"=IF(A{r}=1,{ann_ref}*E{r},0)"
                else:
                    # Recurring benefits: sub_benefit × maturity_factor × discount_factor
                    formula = f"={ann_ref}*B{r}*E{r}"
                c = ws.cell(row=r, column=7+bi, value=formula)
                c.font = Font(name="Arial", color=BLACK, size=9); c.border = _bd()
                c.number_format = "#,##0.0"; c.alignment = Alignment(horizontal="right")

            row += 1

        DATA_END = row - 1

        # Summary rows
        row += 1
        _cell(ws, row, 1, "Total NPV of Benefits", bold=True, bg=C_LIGHT)
        c = ws.cell(row=row, column=6,
                    value=f"=SUM(F{DATA_START}:F{DATA_END})")
        c.font = Font(name="Arial", bold=True, color=C_GREEN, size=10)
        c.number_format = "#,##0.0"; c.border = _bd()
        c.fill = PatternFill("solid", fgColor=C_LIGHT)
        c.alignment = Alignment(horizontal="right")
        # Track overall total so Benefit Breakdown can reference it
        rm.setdefault("spec_benefit_totals", []).append(
            {"measure_name": m["name"], "row": row, "col": 6}
        )
        # Write per-category SUM cells and track their positions
        category_cols = {}
        for bi, (key, _) in enumerate(BEN_TYPES):
            col_idx = 7 + bi
            c = ws.cell(row=row, column=col_idx,
                        value=f"=SUM({get_column_letter(col_idx)}{DATA_START}:{get_column_letter(col_idx)}{DATA_END})")
            c.font = Font(name="Arial", bold=True, color=BLACK, size=10)
            c.number_format = "#,##0.0"; c.border = _bd()
            c.fill = PatternFill("solid", fgColor=C_LIGHT)
            c.alignment = Alignment(horizontal="right")
            category_cols[key] = col_idx
        # Track per-category SUM row/col for Benefit Breakdown to link to
        rm.setdefault("spec_category_totals", []).append({
            "measure_name": m["name"],
            "category_cols": category_cols,
            "sum_row": row,
        })
        row += 1

        _cell(ws, row, 1, "Total Undiscounted Benefit", bold=False, bg=None)
        c = ws.cell(row=row, column=4,
                    value=f"=SUM(D{DATA_START}:D{DATA_END})")
        c.font = Font(name="Arial", color=BLACK, size=10)
        c.number_format = "#,##0.0"; c.border = _bd()
        c.alignment = Alignment(horizontal="right")
        row += 2

    # ── Explanatory benefit formulas (text only) ──────────────────────────────
    _sec(ws, row, 1, "BENEFIT FORMULAS (CONCEPTUAL OVERVIEW)", span=10); row += 1
    if stype == "natural_shading":
        expl_lines = [
            "Avoided mortality (mapped to advanced_benefits.avoided_mortality_npv): vulnerable_population × base_mortality_rate × heat_mortality_factor × heat_reduction_efficiency × maturity_factor(year) × VSL, aggregated over the 50-year horizon and discounted with the global discount rate. As shaded streets reduce ambient temperature and extreme-heat exposure, the probability of heat-related deaths declines along the shaded corridor.",
            "Morbidity savings (advanced_benefits.morbidity_savings_npv): daily_hospital_cost (3,928 NIS) × average_length_of_stay (5.2 days) × heat_attributable_cases_avoided × heat_reduction_efficiency × maturity_factor. Lower peak temperatures reduce emergency department visits and hospital admissions for cardiovascular and respiratory complications, as well as heat exhaustion and dehydration.",
            "Skin cancer prevention (advanced_benefits.skin_cancer_prevention_npv): pedestrians_per_hour × operating_hours (≈8) × UV_reduction_factor (0.75) × skin_cancer_incidence_rate × (treatment_cost + VSLY_loss) × maturity_factor. Continuous shade along boulevards cuts UV exposure for regular users and reduces both direct treatment costs and the loss of statistical life years associated with melanoma and non-melanoma cancers.",
            "Cooling and energy savings: by lowering air and surface temperatures along the street canyon, mature tree canopies reduce sensible heat gains into adjacent buildings, leading to lower cooling energy demand and electricity costs. These energy and comfort benefits are reflected primarily in the annual benefit flow and, where monetised through health and infrastructure channels, contribute to morbidity_savings_npv and the ecosystem-service NPVs.",
            "Ecosystem services — carbon, runoff, air quality, habitat (advanced_benefits.carbon_sequestration_npv, runoff_reduction_npv, air_quality_npv, habitat_creation_npv): carbon is sequestered in woody biomass and soils (~450 NIS/tree/year × tree_density) and valued as avoided social damage from emissions; runoff reduction comes from interception and evapotranspiration, lowering stormwater volumes and treatment needs; air quality improves as leaves capture PM2.5 and other pollutants, reducing health costs; habitat creation reflects 200–500 NIS/m²/year for urban biodiversity corridors supporting birds, insects, and microfauna.",
            "All benefit streams apply an 8-year linear biological maturity ramp (year/8 up to year 8, then 1.0) so that early years reflect immature canopy coverage and later years reflect fully developed shade. Each year’s effective benefits are then discounted over the 50-year project horizon to produce the NPVs reported in the advanced_benefits fields and the Benefit Breakdown sheet."
        ]
    else:
        expl_lines = [
            "Avoided mortality (advanced_benefits.avoided_mortality_npv): catchment_population × base_mortality_rate × heat_mortality_factor × heat_reduction_efficiency × VSL, summed over 50 years and discounted. By cooling indoor environments and rooftop microclimates, green roofs reduce heat stress for residents in the building and, depending on scale, in the surrounding neighbourhood.",
            "Morbidity savings (advanced_benefits.morbidity_savings_npv): daily_hospital_cost (3,928 NIS) × average_length_of_stay (5.2 days) × heat_attributable_cases_avoided × heat_reduction_efficiency (≈0.28). These avoided cases include heat exhaustion, dehydration, and exacerbations of underlying respiratory and cardiovascular conditions driven by heat waves.",
            "Cooling and energy savings: vegetated roofs improve thermal insulation and reduce heat flux into the building, flattening indoor temperature peaks and lowering cooling loads and electricity bills. These energy savings are primarily reflected in the annual benefit stream and may be implicitly captured within morbidity_savings_npv (through avoided health episodes) and ecosystem-service NPVs where energy-related emissions and infrastructure wear are monetised.",
            "Property value uplift (advanced_benefits.property_value_uplift_npv): roof_area_m² × property_value_per_m² × uplift_pct (≈3%), treated as a one-time capitalised benefit near Year 1 and discounted. Buyers and tenants value improved thermal performance, quieter and greener outlooks, and access to rooftop amenities, which is reflected in higher market prices and rents.",
            "Roof longevity extension (advanced_benefits.roof_longevity_npv): (roof_replacement_cost / conventional_roof_lifetime) × roof_longevity_extension_years (≈15), applied as an avoided replacement expenditure at the conventional end-of-life year. The green roof protects the membrane from UV radiation and temperature cycling, slowing degradation and extending the replacement cycle.",
            "Ecosystem services — carbon, runoff, air quality, habitat (advanced_benefits.carbon_sequestration_npv, runoff_reduction_npv, air_quality_npv, habitat_creation_npv): carbon is stored in biomass and substrates (~350 NIS/m²/year × green_roof_area), runoff reduction arises from water retention and delayed release (stormwater_infrastructure_cost_avoided × runoff_reduction_coefficient ≈0.65), air quality improves as vegetation filters particulates and some gaseous pollutants (valued via health_cost_per_unit), and habitat creation (~300 NIS/m²/year × green_roof_area) recognises the value of new ecological niches for insects, birds, and urban flora.",
            "Rainwater management benefits are particularly important for green roofs: by attenuating peak flows and reducing combined sewer overflows, they lower the required capacity and operating costs of drainage and treatment infrastructure. All of these benefit flows are assumed to operate at full capacity from Year 1 (no biological maturity ramp) and are discounted over the 50-year analysis horizon, feeding into the advanced_benefits NPVs and the Benefit Breakdown sheet."
        ]
    for line in expl_lines:
        c = ws.cell(row=row, column=1, value=line)
        c.font = Font(name="Arial", size=9, color="475569")
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
        row += 1

    # ── Column widths ──────────────────────────────────────────────────────────
    _auto_widths(ws)


# ── SHEET 6: BENEFIT BREAKDOWN ────────────────────────────────────────────────
def _benefit_breakdown(wb, data, rm):
    """Show NPV of each benefit category per measure, plus totals and shares."""
    ws = wb.create_sheet("Benefit Breakdown")
    ws.sheet_view.showGridLines = False

    measures = data["measures"]
    n = rm["n"]
    cur = f"{data['currency']} ({data['currency_unit']})"

    _hdr(ws, 1, 1, "BENEFIT BREAKDOWN BY CATEGORY", sz=13, span=n+2)
    _hdr(
        ws,
        2,
        1,
        f"All benefit values in {cur}  |  Uses advanced_benefits NPVs from specialist analysis; cooling and energy savings are captured within the annual benefit streams and allocated across the categories below.",
        bg=C_MID,
        fg="CBD5E1",
        bold=False,
        sz=9,
        span=n+2,
    )

    row = 4
    _sec(ws, row, 1, f"BENEFIT NPVs BY CATEGORY  [{cur}]", span=n+1); row += 1

    # Header row
    _hdr(ws, row, 1, "Benefit Category", bg=C_ACCENT, sz=10)
    for ci, m in enumerate(measures, 2):
        _hdr(ws, row, ci, m["name"], bg=C_ACCENT, sz=10)
    _hdr(ws, row, n+2, "Total Across Measures", bg=C_ACCENT, sz=10)
    row += 1

    BENEFIT_LABELS = [
        ("avoided_mortality_npv",      "Avoided Mortality — NPV"),
        ("morbidity_savings_npv",      "Morbidity Savings — NPV"),
        ("skin_cancer_prevention_npv", "Skin Cancer Prevention — NPV"),
        ("carbon_sequestration_npv",   "Carbon Sequestration — NPV"),
        ("runoff_reduction_npv",       "Runoff Reduction — NPV"),
        ("air_quality_npv",            "Air Quality — NPV"),
        ("habitat_creation_npv",       "Habitat Creation — NPV"),
        ("property_value_uplift_npv",  "Property Value Uplift — NPV"),
        ("roof_longevity_npv",         "Roof Longevity Extension — NPV"),
    ]

    spec_cat = rm.get("spec_category_totals", [])

    first_benefit_row = row
    for key, label in BENEFIT_LABELS:
        # Only show rows that are used by at least one measure
        has_data = (
            any((m.get("advanced_benefits") or {}).get(key) for m in measures)
            or any(
                mi < len(spec_cat) and key in spec_cat[mi].get("category_cols", {})
                for mi in range(len(measures))
            )
        )
        if not has_data:
            continue
        _cell(ws, row, 1, label, bold=False, bg="F0FDF4")
        row_total_cells = []
        for ci, m in enumerate(measures, 2):
            mi = ci - 2
            # Prefer linking to Specialist Detail formula SUM
            formula_val = None
            if mi < len(spec_cat):
                info = spec_cat[mi]
                col_idx = info["category_cols"].get(key)
                if col_idx:
                    formula_val = f"='Specialist Detail'!{get_column_letter(col_idx)}${info['sum_row']}"
            if formula_val:
                c = ws.cell(row=row, column=ci, value=formula_val)
                c.font = Font(name="Arial", color=GREEN_LK, size=10)
            else:
                # Fallback: static value from advanced_benefits
                val = (m.get("advanced_benefits") or {}).get(key, 0) or 0
                c = ws.cell(row=row, column=ci, value=val)
                c.font = Font(name="Arial", color=BLACK, size=10)
            c.number_format = "#,##0.0"; c.border = _bd()
            c.alignment = Alignment(horizontal="right")
            c.fill = PatternFill("solid", fgColor="F0FDF4")
            row_total_cells.append(c.coordinate)
        if row_total_cells:
            total_formula = f"=SUM({row_total_cells[0]}:{row_total_cells[-1]})"
            c = ws.cell(row=row, column=n+2, value=total_formula)
            c.font = Font(name="Arial", bold=False, color=BLACK, size=10)
            c.number_format = "#,##0.0"; c.border = _bd()
            c.alignment = Alignment(horizontal="right")
            c.fill = PatternFill("solid", fgColor="F0FDF4")
        row += 1
    last_benefit_row = row - 1

    # Total PV of benefits per measure
    row += 1
    _cell(ws, row, 1, "Total PV of Benefits (all categories)", bold=True, bg=C_LIGHT)
    for ci in range(2, n+2):
        col_letter = get_column_letter(ci)
        c = ws.cell(
            row=row,
            column=ci,
            value=f"=SUM({col_letter}{first_benefit_row}:{col_letter}{last_benefit_row})",
        )
        c.font = Font(name="Arial", bold=True, color=C_GREEN, size=10)
        c.number_format = "#,##0.0"; c.border = _bd()
        c.fill = PatternFill("solid", fgColor=C_LIGHT)
        c.alignment = Alignment(horizontal="right")
    row += 2

    # Share of total benefits by category (per measure)
    _sec(ws, row, 1, "SHARE OF TOTAL BENEFITS BY CATEGORY  (per measure)", span=n+1); row += 1
    _hdr(ws, row, 1, "Benefit Category", bg=C_ACCENT, sz=10)
    for ci, m in enumerate(measures, 2):
        _hdr(ws, row, ci, m["name"], bg=C_ACCENT, sz=10)
    row += 1

    total_row = last_benefit_row + 2  # row where total PV of benefits per measure was written
    share_start_row = row
    for r in range(first_benefit_row, last_benefit_row + 1):
        label = ws.cell(row=r, column=1).value
        _cell(ws, row, 1, label, bold=False)
        for ci in range(2, n+2):
            num = f"{get_column_letter(ci)}{r}"
            den = f"{get_column_letter(ci)}{total_row}"
            c = ws.cell(row=row, column=ci, value=f"=IF({den}>0,{num}/{den},0)")
            c.font = Font(name="Arial", color=BLACK, size=10)
            c.number_format = "0.0%"; c.border = _bd()
            c.alignment = Alignment(horizontal="right")
        row += 1

    row += 2
    _sec(ws, row, 1, "LINKS TO YEAR-BY-YEAR BENEFITS (Specialist Detail)", span=n+1); row += 1
    _hdr(ws, row, 1, "Measure", bg=C_ACCENT, sz=10)
    _hdr(ws, row, 2, "Total NPV of Benefits (Specialist Detail, col F)", bg=C_ACCENT, sz=10)
    _hdr(ws, row, 3, "Total PV of Benefits (from table above)", bg=C_ACCENT, sz=10)
    row += 1

    spec_totals = rm.get("spec_benefit_totals", [])
    for mi, m in enumerate(measures):
        _cell(ws, row, 1, m["name"], bold=True)
        # If we have recorded a Specialist Detail total row for this measure, link to it
        link_cell_val = ""
        if mi < len(spec_totals):
            info = spec_totals[mi]
            spec_row = info["row"]
            spec_col = info["col"]
            spec_coord = f"$F${spec_row}" if spec_col == 6 else f"{get_column_letter(spec_col)}{spec_row}"
            link_cell_val = f"='Specialist Detail'!{spec_coord}"
        c = ws.cell(row=row, column=2, value=link_cell_val or None)
        if link_cell_val:
            c.font = Font(name="Arial", color=GREEN_LK, size=10)
            c.number_format = "#,##0.0"; c.border = _bd()
            c.alignment = Alignment(horizontal="right")

        total_col_letter = get_column_letter(mi + 2)
        c2 = ws.cell(row=row, column=3, value=f"={total_col_letter}{total_row}")
        c2.font = Font(name="Arial", color=BLACK, size=10)
        c2.number_format = "#,##0.0"; c2.border = _bd()
        c2.alignment = Alignment(horizontal="right")

        row += 1

    _auto_widths(ws)


# ── SHEET 7: BENEFIT DETAIL ───────────────────────────────────────────────────
# Formula engine: builds live Excel formulas from structured benefit_components JSON
# Each formula type maps to a builder function that writes labelled parameter rows
# and returns the Excel cell reference of the computed annual benefit value.

_BENEFIT_SOURCES = {
    "avoided_mortality":        "VSL: OECD (2005) baseline, CPI/PPP-adjusted; Heat-mortality factor: Gasparrini et al. (2017) Lancet",
    "morbidity_savings":        "Hospitalization cost: Israeli Health Ministry (2024); Average LOS: clinical literature; Cases-avoided: CDD model",
    "skin_cancer_prevention":   "UV-cancer incidence: WHO/ICO Global Cancer Observatory; Treatment costs: Ministry of Health; UV reduction: solar geometry literature",
    "carbon_sequestration":     "Carbon price: Israeli carbon market reference (2024); Sequestration rate: urban forestry literature (Nowak et al. 2002)",
    "runoff_reduction":         "Runoff coefficient: EPA Stormwater BMP Guide; Infrastructure cost avoided: local municipality data",
    "air_quality":              "PM2.5 health cost: WHO Air Quality Guidelines (2021); PM removal rates: vegetation science literature",
    "habitat_creation":         "Biodiversity value: TEEB (2010); 200-500 NIS/m2/yr range from Israeli urban ecology studies",
    "property_value_uplift":    "Hedonic pricing meta-analysis: Fuerst & McAllister (2011); 3% uplift on adjacent property value per m2 green roof",
    "roof_longevity":           "Lifespan extension: Berghage et al. (2009) GRHC; Conventional roof replacement cost: industry data",
    "thermal_comfort":          "Thermal comfort monetisation: ASHRAE 55 standard; avoided cooling cost literature",
    "flood_protection":         "Avoided damage: national flood risk databases; infrastructure replacement cost method",
    "tourism":                  "Visitor spending: local tourism authority data; contingent valuation literature",
    "energy_savings":           "Energy reduction per m2 shade/green roof: building physics simulation literature",
    "default":                  "Literature estimate - see data_source field in Inputs sheet for measure-specific citations",
}

_BENEFIT_CALCS = {
    "avoided_mortality":      "Population x base mortality rate x heat-mortality factor x heat reduction efficiency x maturity factor x VSL",
    "morbidity_savings":      "Hospitalization cost (3,928 NIS/day) x avg LOS (5.2 days) x heat-attributable cases avoided x efficiency x maturity",
    "skin_cancer_prevention": "Pedestrians/hr x operating hours (8) x UV reduction (0.75) x incidence rate x (treatment cost + VSLY) x maturity",
    "carbon_sequestration":   "Carbon price (NIS/unit) x sequestration rate x number of trees or roof area",
    "runoff_reduction":       "Runoff reduction coefficient x stormwater infrastructure cost avoided per m3",
    "air_quality":            "PM2.5 removed (kg/yr) x health cost per kg x affected population",
    "habitat_creation":       "Green area (m2) x biodiversity unit value (200-500 NIS/m2/yr)",
    "property_value_uplift":  "Roof area (m2) x property value per m2 x uplift fraction (0.03) - one-time Year 1 benefit",
    "roof_longevity":         "(Roof replacement cost / conventional roof life) x longevity extension years (15) - lump sum at end-of-life",
    "thermal_comfort":        "Avoided cooling energy (kWh/yr) x electricity tariff + comfort value to occupants",
    "flood_protection":       "Expected annual damage avoided x probability of flood event",
    "tourism":                "Additional visitors x average spending per visitor x attribution fraction",
    "energy_savings":         "Energy reduction (kWh/m2/yr) x green area x electricity tariff",
    "default":                "Net present value of annual benefit stream discounted at project discount rate",
}

_BENEFIT_ENDO_EXOG = {
    "avoided_mortality":      ("Endogenic: population count, project area\nExogenic: VSL (OECD), heat-mortality factor, CDD baseline", "Exogenic"),
    "morbidity_savings":      ("Endogenic: project area, efficiency factor\nExogenic: hospitalization cost, LOS, CDD incidence rate", "Exogenic"),
    "skin_cancer_prevention": ("Endogenic: pedestrian counts, operating hours\nExogenic: UV reduction rate, incidence, treatment cost, VSLY", "Exogenic"),
    "carbon_sequestration":   ("Endogenic: number of trees / roof area\nExogenic: carbon price, sequestration rate per tree", "Exogenic"),
    "runoff_reduction":       ("Endogenic: catchment area, surface type\nExogenic: runoff coefficients, infrastructure cost/m3", "Exogenic"),
    "air_quality":            ("Endogenic: green area, local traffic levels\nExogenic: PM removal rate, health cost per kg PM2.5", "Exogenic"),
    "habitat_creation":       ("Endogenic: green area (m2)\nExogenic: biodiversity unit value (literature range)", "Exogenic"),
    "property_value_uplift":  ("Endogenic: roof area, local property values\nExogenic: uplift fraction from hedonic pricing literature", "Mixed"),
    "roof_longevity":         ("Endogenic: roof area, local replacement cost\nExogenic: longevity extension years (literature)", "Mixed"),
    "thermal_comfort":        ("Endogenic: building area, occupancy\nExogenic: energy tariff, thermal comfort value", "Mixed"),
    "flood_protection":       ("Endogenic: protected area, asset values\nExogenic: damage curves, flood return period", "Mixed"),
    "tourism":                ("Endogenic: project location, visitor estimates\nExogenic: average visitor spending, attribution factor", "Mixed"),
    "energy_savings":         ("Endogenic: green area, building type\nExogenic: energy reduction rate (simulation), electricity tariff", "Mixed"),
    "default":                ("Endogenic: project-specific inputs\nExogenic: literature unit values and multipliers", "Mixed"),
}


def _benefit_detail(wb, data, rm=None):
    """
    Benefit Detail sheet — formula-engine version.
    For each measure, each benefit component is computed via a live Excel formula
    built from individual parameters (population, VSL, efficiency, etc.).
    Every parameter shows its value, type (endogenous/exogenous), and source citation.
    """
    ws = wb.create_sheet("Benefit Detail")
    ws.sheet_view.showGridLines = False

    measures  = data["measures"]
    cur       = data.get("currency", "EUR")
    cur_unit  = data.get("currency_unit", "millions")
    dr        = data.get("discount_rate", 0.035)
    horizon   = data.get("time_horizon", 30)
    SPAN      = 6

    # ── sheet-local helpers ────────────────────────────────────────────────────
    def _title(r, txt):
        c = ws.cell(row=r, column=1, value=txt)
        c.font      = Font(name="Arial", bold=True, color=C_WHITE, size=11)
        c.fill      = PatternFill("solid", fgColor=C_DARK)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border    = _bd()
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=SPAN)
        ws.row_dimensions[r].height = 22

    def _sub(r, txt, bg=C_MID):
        c = ws.cell(row=r, column=1, value=txt)
        c.font      = Font(name="Arial", bold=True, color=C_WHITE, size=10)
        c.fill      = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border    = _bd()
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=SPAN)
        ws.row_dimensions[r].height = 18

    def _col_hdr(r):
        for ci, h in enumerate(
            ["Parameter", "Value", "Unit", "Type", "Source / Citation", "Notes"], 1
        ):
            c = ws.cell(row=r, column=ci, value=h)
            c.font      = Font(name="Arial", bold=True, color=C_WHITE, size=9)
            c.fill      = PatternFill("solid", fgColor=C_ACCENT)
            c.border    = _bd()
            c.alignment = Alignment(horizontal="center", vertical="center")

    def _prow(r, label, value, unit, dtype, source, note="",
              is_endo=False, is_exog=False, is_comp=False, is_glob=False, fmt=None):
        """One parameter row."""
        if is_endo:  bg, fg = C_ENDO_BG, C_ENDO_FG
        elif is_exog: bg, fg = C_EXOG_BG, C_EXOG_FG
        elif is_comp: bg, fg = C_COMP_BG, C_COMP_FG
        elif is_glob: bg, fg = C_GLOB_BG, C_GLOB_FG
        else:         bg, fg = "F8FAFC",  "475569"
        type_label = ("🔵 Local input" if is_endo
                      else "🟠 Literature" if is_exog
                      else "🟢 Computed"   if is_comp
                      else "🟡 Global"     if is_glob
                      else "")
        vals = [label, value, unit, type_label, source, note]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=ci, value=v)
            c.font      = Font(name="Arial", size=9, color=fg,
                               bold=(ci == 1 and (is_comp or is_glob)))
            c.fill      = PatternFill("solid", fgColor=bg)
            c.border    = _bd()
            c.alignment = Alignment(wrap_text=True, vertical="center",
                                    horizontal="right" if ci == 2 else "left")
            if ci == 2 and fmt:
                c.number_format = fmt
        ws.row_dimensions[r].height = 18

    def _result(r, label, formula, source):
        """► highlighted result row — contains a live Excel formula."""
        vals = [f"► {label}", formula, f"{cur}M / year", "🟢 Computed", source, "KEY OUTPUT → drives NPV"]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=ci, value=v)
            c.font      = Font(name="Arial", bold=True, size=9, color=C_COMP_FG)
            c.fill      = PatternFill("solid", fgColor=C_COMP_BG)
            c.border    = _bd()
            c.alignment = Alignment(wrap_text=True, vertical="center",
                                    horizontal="right" if ci == 2 else "left")
            if ci == 2:
                c.number_format = "#,##0.000"
        ws.row_dimensions[r].height = 20
        return ws.cell(row=r, column=2)   # return the formula cell ref

    def _sum_result(r, label, component_refs, source):
        """► sum row that adds all component result cells."""
        if component_refs:
            formula = "=" + "+".join(component_refs)
        else:
            formula = 0
        return _result(r, label, formula, source)

    def _blank(r):
        ws.row_dimensions[r].height = 7

    # ── Formula builders ───────────────────────────────────────────────────────
    # Each returns (component_cell_address, row_after)

    def _build_avoided_mortality(comp, start_row):
        """avoided_mortality: pop × mort_rate × hmf × efficiency × VSL / 1e6"""
        r = start_row
        _sub(r, f"  ⬛ {comp.get('name','Avoided Mortality')}  —  Formula: Population × Mortality Rate × Heat-Mortality Factor × Efficiency × VSL", bg="1E3A5F"); r+=1
        _col_hdr(r); r+=1

        pop   = comp.get("population_at_risk", 0)
        mrate = comp.get("mortality_rate", 0.010)
        hmf   = comp.get("heat_mortality_factor", 0.00083)
        eff   = comp.get("heat_reduction_efficiency", 0.28)
        vsl   = comp.get("vsl", 3800000)
        # Normalise VSL: Claude sometimes returns 3.8 (millions) instead of 3,800,000
        # If VSL < 1000 it was almost certainly expressed in millions — convert to units
        if isinstance(vsl, (int, float)) and 0 < vsl < 1000:
            vsl = vsl * 1_000_000

        r_pop   = r; _prow(r, "Population at Risk (persons)",            pop,   "persons",  "", comp.get("population_at_risk_source","User provided"),          is_endo=True,  fmt="#,##0"); r+=1
        r_mrate = r; _prow(r, "Base Annual Mortality Rate",              mrate, "deaths/person/yr", "", comp.get("mortality_rate_source","Ministry of Health"),   is_exog=True,  fmt="0.0000"); r+=1
        r_hmf   = r; _prow(r, "Heat-Attributable Mortality Fraction (annual, fraction of deaths)", hmf, "fraction of annual deaths", "", comp.get("heat_mortality_factor_source","Gasparrini et al. 2017"), is_exog=True, fmt="0.000000"); r+=1
        r_eff   = r; _prow(r, "Heat Reduction Efficiency of this measure", eff, "fraction 0–1", "", comp.get("heat_reduction_efficiency_source","Literature"),    is_exog=True,  fmt="0.00"); r+=1
        r_vsl   = r; _prow(r, f"VSL — Value of Statistical Life ({cur})", vsl,  cur,          "", comp.get("vsl_source","OECD 2012 meta-analysis"),              is_exog=True,  fmt="#,##0"); r+=1

        # intermediate steps
        _prow(r, "  Deaths avoided per year (before VSL)",
              f"=B{r_pop}*B{r_mrate}*B{r_hmf}*B{r_eff}", "deaths/yr", "", "Formula derivation", is_comp=True, fmt="0.0000"); r_deaths=r-1; r+=1

        formula = f"=B{r_pop}*B{r_mrate}*B{r_hmf}*B{r_eff}*B{r_vsl}/1000000"
        cell = _result(r, comp.get("name","Avoided Mortality"), formula,
                       "Population × Mortality Rate × HMF × Efficiency × VSL ÷ 1,000,000"); r+=1
        return cell.coordinate, r

    def _build_energy_savings(comp, start_row):
        """energy_savings: area × kwh_reduction × tariff / 1e6"""
        r = start_row
        _sub(r, f"  ⬛ {comp.get('name','Energy Savings')}  —  Formula: Area × Energy Reduction (kWh/m²) × Electricity Tariff", bg="1E3A5F"); r+=1
        _col_hdr(r); r+=1

        area    = comp.get("area_m2", 0)
        kwh     = comp.get("energy_reduction_kwh_m2", 25)
        tariff  = comp.get("electricity_tariff", 0.14)

        r_area   = r; _prow(r, "Project Area (m²)",                    area,   "m²",          "", comp.get("area_m2_source","User provided"),                is_endo=True,  fmt="#,##0"); r+=1
        r_kwh    = r; _prow(r, "Energy Reduction (kWh/m²/year)",       kwh,    "kWh/m²/yr",   "", comp.get("energy_reduction_kwh_m2_source","Literature"),   is_exog=True,  fmt="0.0"); r+=1
        r_tariff = r; _prow(r, f"Electricity Tariff ({cur}/kWh)",      tariff, f"{cur}/kWh",  "", comp.get("electricity_tariff_source","Local utility"),     is_endo=True,  fmt="0.000"); r+=1

        formula = f"=B{r_area}*B{r_kwh}*B{r_tariff}/1000000"
        cell = _result(r, comp.get("name","Energy Savings"), formula,
                       "Area × Energy Reduction × Tariff ÷ 1,000,000"); r+=1
        return cell.coordinate, r

    def _build_morbidity_savings(comp, start_row):
        """morbidity_savings: cases × hosp_cost × los / 1e6"""
        r = start_row
        _sub(r, f"  ⬛ {comp.get('name','Morbidity Savings')}  —  Formula: Cases Avoided × Hospitalization Cost × Length of Stay", bg="1E3A5F"); r+=1
        _col_hdr(r); r+=1

        cases   = comp.get("cases_avoided_per_year", 0)
        hcost   = comp.get("hospitalization_cost", 3928)
        los     = comp.get("avg_length_of_stay_days", 5.2)

        r_cases = r; _prow(r, "Heat-Attributable Cases Avoided / Year", cases,  "cases/yr",    "", comp.get("cases_source","User estimate / literature"),      is_endo=True,  fmt="#,##0"); r+=1
        r_hcost = r; _prow(r, f"Hospitalization Cost ({cur}/day)",      hcost,  f"{cur}/day",  "", comp.get("hospitalization_cost_source","Ministry of Health"), is_exog=True, fmt="#,##0"); r+=1
        r_los   = r; _prow(r, "Average Length of Stay (days)",          los,    "days",        "", comp.get("avg_length_of_stay_days_source","Clinical literature"), is_exog=True, fmt="0.0"); r+=1

        formula = f"=B{r_cases}*B{r_hcost}*B{r_los}/1000000"
        cell = _result(r, comp.get("name","Morbidity Savings"), formula,
                       "Cases Avoided × Hosp. Cost × Length of Stay ÷ 1,000,000"); r+=1
        return cell.coordinate, r

    def _build_property_value_uplift(comp, start_row):
        """property_value_uplift: area × value_per_m2 × uplift_fraction / 1e6"""
        r = start_row
        _sub(r, f"  ⬛ {comp.get('name','Property Value Uplift')}  —  Formula: Area × Property Value × Uplift Fraction", bg="1E3A5F"); r+=1
        _col_hdr(r); r+=1

        area    = comp.get("affected_area_m2", 0)
        val_m2  = comp.get("property_value_per_m2", 0)
        uplift  = comp.get("uplift_fraction", 0.03)

        r_area  = r; _prow(r, "Affected Area (m²)",                    area,   "m²",          "", comp.get("area_m2_source","User provided"),             is_endo=True,  fmt="#,##0"); r+=1
        r_val   = r; _prow(r, f"Property Value per m² ({cur})",        val_m2, f"{cur}/m²",   "", comp.get("property_value_per_m2_source","Local market"), is_endo=True,  fmt="#,##0"); r+=1
        r_upl   = r; _prow(r, "Uplift Fraction",                       uplift, "fraction",    "", comp.get("uplift_fraction_source","Fuerst & McAllister 2011"), is_exog=True, fmt="0.000"); r+=1

        # Property uplift is a one-time capital gain → annualise over measure lifetime
        _lifetime = comp.get("lifetime_years", 30)
        formula = f"=B{r_area}*B{r_val}*B{r_upl}/1000000/{_lifetime}"
        _prow(r, "Measure Lifetime (for annualisation)", _lifetime, "years", "",
              "Divides one-time capital gain into annual equivalent",
              "Property uplift is a one-time gain — annualised for CBA flow",
              is_exog=True, fmt="0"); r+=1
        cell = _result(r, comp.get("name","Property Value Uplift"), formula,
                       f"Area × Value/m² × Uplift Fraction ÷ 1,000,000 ÷ {_lifetime} yrs (annualised)"); r+=1
        return cell.coordinate, r

    def _build_generic(comp, start_row):
        """generic_annual: pre-computed value (fallback)."""
        r = start_row
        _sub(r, f"  ⬛ {comp.get('name','Benefit Component')}  —  Annual value (literature estimate)", bg="334155"); r+=1
        _col_hdr(r); r+=1

        val    = comp.get("annual_value", comp.get("value", 0))
        source = comp.get("source", comp.get("annual_value_source","Literature estimate"))

        r_val  = r; _prow(r, comp.get("name","Annual Benefit"), val,
                          f"{cur}M/yr", "", source,
                          "🟠 Pre-computed literature value — no local formula available",
                          is_exog=True, fmt="#,##0.000"); r+=1
        formula = f"=B{r_val}"
        cell = _result(r, comp.get("name","Benefit"), formula, source); r+=1
        return cell.coordinate, r

    FORMULA_BUILDERS = {
        "avoided_mortality":      _build_avoided_mortality,
        "energy_savings":         _build_energy_savings,
        "morbidity_savings":      _build_morbidity_savings,
        "property_value_uplift":  _build_property_value_uplift,
        "generic_annual":         _build_generic,
    }

    # ── Page header ────────────────────────────────────────────────────────────
    _hdr(ws, 1, 1, "BENEFIT CALCULATION AUDIT TRAIL — LIVE EXCEL FORMULAS PER MEASURE", sz=12, span=SPAN)
    _hdr(ws, 2, 1,
         f"All monetary values in {cur} ({cur_unit})  |  "
         "Change any 🔵 blue parameter → all results update automatically",
         bg=C_MID, fg="CBD5E1", bold=False, sz=9, span=SPAN)
    _add_legend(ws, 3, 1, span=SPAN)
    row = 10

    # Global params
    _sub(row, "GLOBAL PARAMETERS  (apply to all measures)", bg="334155"); row+=1
    _col_hdr(row); row+=1
    _prow(row, "Discount Rate", dr, "%", "", "User input — Inputs sheet", is_glob=True, fmt="0.0%"); row+=1
    _prow(row, "Time Horizon", horizon, "years", "", "User input — Inputs sheet", is_glob=True, fmt="0"); row+=1
    _blank(row); row+=1

    # Annuity formula explanation
    _sub(row, "PRESENT VALUE FORMULA  →  PV = Annual Benefit × [(1−(1+r)^−n) / r]", bg="334155"); row+=1
    for txt in [
        "Annuity Factor = (1−(1+r)^−n) / r   →   converts a constant annual stream to present value",
        "NPV = PV(Benefits) − CAPEX − PV(OPEX)   →   total economic value created",
        "BCR = PV(Benefits) / [CAPEX + PV(OPEX)]   →   return per unit invested  (>1.0 = viable,  >1.5 = recommended)",
    ]:
        c = ws.cell(row=row, column=1, value=txt)
        c.font = Font(name="Arial", size=9, color=C_COMP_FG, bold=True)
        c.fill = PatternFill("solid", fgColor=C_COMP_BG)
        c.border = _bd()
        c.alignment = Alignment(wrap_text=True, vertical="center")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=SPAN)
        ws.row_dimensions[row].height = 16
        row+=1
    _blank(row); row+=1

    # ── Per-measure blocks ─────────────────────────────────────────────────────
    summary_rows = []   # (name, capex, pv_ben_ref, pv_cost_ref) for final table

    for m_idx, m in enumerate(measures):
        mname    = m.get("name", f"Measure {m_idx+1}")
        capex    = m.get("capex", 0)
        opex     = m.get("annual_opex", 0)
        lifetime = m.get("lifetime_years", horizon)
        desc     = m.get("description", "")
        cobens   = m.get("co_benefits", "")
        feas     = m.get("feasibility", "")
        uncert   = m.get("uncertainty", "")

        _title(row, f"MEASURE {m_idx+1} OF {len(measures)}: {mname.upper()}"); row+=1

        # description
        c = ws.cell(row=row, column=1, value=desc)
        c.font = Font(name="Arial", size=9, color="475569", italic=True)
        c.border = _bd()
        c.alignment = Alignment(wrap_text=True, vertical="center")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=SPAN)
        ws.row_dimensions[row].height = max(18, len(desc)//8*5+14)
        row+=1

        # ── COSTS ─────────────────────────────────────────────────────────────
        _sub(row, "  COSTS", bg="1E3A5F"); row+=1
        _col_hdr(row); row+=1

        r_capex  = row
        _prow(row, "Capital Cost (CAPEX)", capex, f"{cur}M", "",
              m.get("capex_source","User provided"), "One-time investment",
              is_endo=True, fmt="#,##0.000"); row+=1
        r_opex   = row
        _prow(row, "Annual O&M (OPEX)", opex, f"{cur}M/yr", "",
              m.get("opex_source","User provided"), "Recurring annual cost",
              is_endo=True, fmt="#,##0.000"); row+=1
        r_life   = row
        _prow(row, "Measure Lifetime", lifetime, "years", "",
              "Project engineering life", "", is_endo=True, fmt="0"); row+=1

        # annuity factor + PV costs (computed rows)
        r_af_c   = row
        _prow(row, "Annuity Factor (OPEX stream)", None,
              "dimensionless", "",
              "Standard annuity formula",
              f"=(1−(1+{dr})^−{lifetime})/{dr}",
              is_comp=True, fmt="0.0000"); row+=1
        ws.cell(row=r_af_c, column=2).value = f"=(1-(1+{dr})^(-{lifetime}))/{dr}"
        ws.cell(row=r_af_c, column=2).number_format = "0.0000"

        r_pv_cost = row
        _prow(row, "► PV of Total Costs", None, f"{cur}M", "",
              "CAPEX + OPEX × Annuity Factor",
              "Denominator of BCR",
              is_comp=True, fmt="#,##0.000"); row+=1
        ws.cell(row=r_pv_cost, column=2).value = f"=B{r_capex}+B{r_opex}*B{r_af_c}"
        ws.cell(row=r_pv_cost, column=2).font  = Font(name="Arial", bold=True, color=C_COMP_FG, size=9)
        ws.cell(row=r_pv_cost, column=2).fill  = PatternFill("solid", fgColor=C_COMP_BG)
        ws.cell(row=r_pv_cost, column=2).number_format = "#,##0.000"
        _blank(row); row+=1

        # ── BENEFITS — formula engine ──────────────────────────────────────────
        _sub(row, "  BENEFITS — STEP-BY-STEP FORMULA DERIVATION", bg="14532D"); row+=1

        components = m.get("benefit_components") or []

        # back-compat: if benefit_components is a dict (old format), convert
        if isinstance(components, dict):
            components = [
                {"name": k.replace("_"," ").title(), "type":"generic_annual", "annual_value": v}
                for k, v in components.items()
                if isinstance(v, (int, float))
            ]

        # if still empty, create one generic from annual_benefit
        if not components:
            ab = m.get("annual_benefit", 0)
            components = [{"name": "Annual Benefit", "type": "generic_annual",
                           "annual_value": ab,
                           "source": m.get("data_source","Literature estimate")}]

        component_result_refs = []

        for comp in components:
            ftype   = comp.get("type", "generic_annual")
            builder = FORMULA_BUILDERS.get(ftype, _build_generic)
            cell_ref, row = builder(comp, row)
            component_result_refs.append(cell_ref)
            _blank(row); row+=1

        # Total annual benefit = sum of all components
        r_total_ben = row
        if len(component_result_refs) == 1:
            total_formula = f"={component_result_refs[0]}"
        else:
            total_formula = "=" + "+".join(component_result_refs)

        c = ws.cell(row=row, column=1, value="► TOTAL ANNUAL BENEFIT (sum of all components above)")
        c.font  = Font(name="Arial", bold=True, color=C_COMP_FG, size=10)
        c.fill  = PatternFill("solid", fgColor=C_COMP_BG)
        c.border= _bd()
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=1)

        c2 = ws.cell(row=row, column=2, value=total_formula)
        c2.font         = Font(name="Arial", bold=True, color=C_COMP_FG, size=10)
        c2.fill         = PatternFill("solid", fgColor=C_COMP_BG)
        c2.border       = _bd()
        c2.number_format= "#,##0.000"
        c2.alignment    = Alignment(horizontal="right", vertical="center")

        for ci, txt in enumerate([f"{cur}M / year","🟢 Computed","= sum of components above","→ feeds NPV & BCR"], 3):
            c3 = ws.cell(row=row, column=ci, value=txt)
            c3.font  = Font(name="Arial", bold=True, color=C_COMP_FG, size=9)
            c3.fill  = PatternFill("solid", fgColor=C_COMP_BG)
            c3.border= _bd()
            c3.alignment = Alignment(vertical="center")
        ws.row_dimensions[row].height = 22
        r_total_ben = row; row+=1

        # Annuity factor + PV of benefits
        r_af_b = row
        _prow(row, "Annuity Factor (benefit stream)", None, "dimensionless", "",
              "Standard annuity formula", f"=(1−(1+{dr})^−{lifetime})/{dr}",
              is_comp=True, fmt="0.0000"); row+=1
        ws.cell(row=r_af_b, column=2).value         = f"=(1-(1+{dr})^(-{lifetime}))/{dr}"
        ws.cell(row=r_af_b, column=2).number_format = "0.0000"

        r_pv_ben = row
        _prow(row, "► PV of Total Benefits", None, f"{cur}M", "",
              "Total Annual Benefit × Annuity Factor", "Numerator of BCR",
              is_comp=True, fmt="#,##0.000"); row+=1
        ws.cell(row=r_pv_ben, column=2).value         = f"=B{r_total_ben}*B{r_af_b}"
        ws.cell(row=r_pv_ben, column=2).font          = Font(name="Arial", bold=True, color=C_COMP_FG, size=9)
        ws.cell(row=r_pv_ben, column=2).fill          = PatternFill("solid", fgColor=C_COMP_BG)
        ws.cell(row=r_pv_ben, column=2).number_format = "#,##0.000"
        _blank(row); row+=1

        # ── RESULTS ───────────────────────────────────────────────────────────
        _sub(row, "  FINANCIAL RESULTS", bg=C_GREEN); row+=1

        r_npv = row
        _prow(row, "► NPV", None, f"{cur}M", "",
              "PV Benefits − PV Costs", "Positive = economically viable",
              is_comp=True, fmt="#,##0.000"); row+=1
        ws.cell(row=r_npv, column=2).value         = f"=B{r_pv_ben}-B{r_pv_cost}"
        ws.cell(row=r_npv, column=2).number_format = "#,##0.000"

        r_bcr = row
        _prow(row, "► BCR", None, "ratio", "",
              "PV Benefits / PV Costs", "BCR > 1.5 = Recommended",
              is_comp=True, fmt="0.00"); row+=1
        ws.cell(row=r_bcr, column=2).value         = f"=IF(B{r_pv_cost}>0,B{r_pv_ben}/B{r_pv_cost},0)"
        ws.cell(row=r_bcr, column=2).number_format = "0.00"

        r_pay = row
        _prow(row, "► Simple Payback (years)", None, "years", "",
              "CAPEX / (Annual Benefit − OPEX)", "",
              is_comp=True, fmt='0.0'); row+=1
        ws.cell(row=r_pay, column=2).value = (
            f'=IF((B{r_total_ben}-B{r_opex})>0,B{r_capex}/(B{r_total_ben}-B{r_opex}),"N/A")'
        )

        # Co-benefits
        if cobens:
            _prow(row, "Co-benefits (not in BCR)", cobens, "", "",
                  "Expert assessment / literature", "Important for policy narrative"); row+=1

        _prow(row, "Feasibility",       feas,   "","","Expert assessment",""); row+=1
        _prow(row, "Uncertainty Level", uncert, "","","Data quality assessment",""); row+=1

        _blank(row); row+=1
        _blank(row); row+=1

        summary_rows.append({
            "name":          mname,
            "capex":         capex,
            "pv_ben_ref":    f"B{r_pv_ben}",
            "pv_cost_ref":   f"B{r_pv_cost}",
            "npv_ref":       f"B{r_npv}",
            "bcr_ref":       f"B{r_bcr}",
            "total_ben_row": r_total_ben,   # row of TOTAL ANNUAL BENEFIT in this sheet
        })

    # ── Summary comparison table ───────────────────────────────────────────────
    _title(row, "MEASURES COMPARISON — BCR RANKING (live formulas — auto-updates)"); row+=1
    for ci, h in enumerate(
        ["#","Measure","CAPEX","PV Benefits","PV Costs","NPV","BCR","Viable?"], 1
    ):
        c = ws.cell(row=row, column=ci, value=h)
        c.font      = Font(name="Arial", bold=True, color=C_WHITE, size=9)
        c.fill      = PatternFill("solid", fgColor=C_ACCENT)
        c.border    = _bd()
        c.alignment = Alignment(horizontal="center", vertical="center")
    row+=1

    for rank, sr in enumerate(summary_rows, 1):
        bcr_ref   = sr["bcr_ref"]
        bg = C_COMP_BG
        fg = C_COMP_FG
        vals = [
            rank,
            sr["name"],
            sr["capex"],
            f"={sr['pv_ben_ref']}",
            f"={sr['pv_cost_ref']}",
            f"={sr['npv_ref']}",
            f"={sr['bcr_ref']}",
            f'=IF({bcr_ref}>=1.5,"✓ Recommended",IF({bcr_ref}>=1.0,"○ Consider","✗ Review"))',
        ]
        fmts = [None, None, "#,##0.00","#,##0.00","#,##0.00","#,##0.00","0.00", None]
        for ci, (v, fmt) in enumerate(zip(vals, fmts), 1):
            c = ws.cell(row=row, column=ci, value=v)
            c.font      = Font(name="Arial", size=9, color=fg, bold=(ci in (2,7)))
            c.fill      = PatternFill("solid", fgColor=bg)
            c.border    = _bd()
            c.alignment = Alignment(
                horizontal="right" if ci in (1,3,4,5,6,7) else "left",
                vertical="center"
            )
            if fmt:
                c.number_format = fmt
        ws.row_dimensions[row].height = 18
        row+=1

    # Column widths
    for col, w in {1:30, 2:14, 3:10, 4:38, 5:44, 6:28}.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    # Return per-measure total row addresses for Inputs linkage
    # Key: measure index (0-based), Value: "'Benefit Detail'!B<row>"
    return {
        idx: f"'Benefit Detail'!B{sr['total_ben_row']}"
        for idx, sr in enumerate(summary_rows)
    }
