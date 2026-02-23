from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

C_DARK   = "1A1A2E"
C_MID    = "16213E"
C_ACCENT = "3B82F6"
C_GREEN  = "059669"
C_RED    = "DC2626"
C_AMBER  = "D97706"
C_LIGHT  = "EFF6FF"
C_WHITE  = "FFFFFF"
C_BORDER = "E2E8F0"
BLUE     = "0000FF"
BLACK    = "000000"
GREEN_LK = "008000"


def _bd():
    s = Side(style="thin", color=C_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)

def _hdr(ws, r, c, v, bg=C_DARK, fg=C_WHITE, bold=True, sz=11, wrap=False, span=1):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = Font(name="Arial", bold=bold, color=fg, size=sz)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
    if span > 1:
        ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c+span-1)
    return cell

def _cell(ws, r, c, v, bold=False, color=BLACK, fmt=None, bg=None, align="left"):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = Font(name="Arial", bold=bold, color=color, size=10)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border = _bd()
    if fmt: cell.number_format = fmt
    if bg:  cell.fill = PatternFill("solid", fgColor=bg)
    return cell

def _sec(ws, r, c, label, span=6):
    cell = ws.cell(row=r, column=c, value=label)
    cell.font = Font(name="Arial", bold=True, color=C_WHITE, size=10)
    cell.fill = PatternFill("solid", fgColor=C_MID)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c+span-1)
    ws.row_dimensions[r].height = 18

def _widths(ws, d):
    for col, w in d.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def build_excel(data: dict, path: str):
    wb = Workbook()
    rm = _inputs(wb, data)
    _results(wb, data, rm)
    _sensitivity(wb, data, rm)
    _summary(wb, data, rm)
    wb.save(path)


# ── SHEET 1: INPUTS ────────────────────────────────────────────────────────────
def _inputs(wb, data):
    ws = wb.active
    ws.title = "Inputs"
    ws.sheet_view.showGridLines = False

    measures = data["measures"]
    n = len(measures)
    cur = f"{data['currency']} ({data['currency_unit']})"

    _hdr(ws, 1, 1, f"CLIMATE ADAPTATION CBA — {data['problem_title'].upper()}", sz=13, span=n+4)
    _hdr(ws, 2, 1, data["problem_summary"], bg=C_MID, fg="CBD5E1", bold=False, sz=9, wrap=True, span=n+4)
    ws.row_dimensions[2].height = 36

    row = 4
    _sec(ws, row, 1, "GLOBAL PARAMETERS", span=4); row += 1

    _cell(ws, row, 1, "Discount Rate", bold=True)
    c = ws.cell(row=row, column=2, value=data["discount_rate"])
    c.font = Font(name="Arial", bold=True, color=BLUE, size=10)
    c.number_format = "0.00%"; c.border = _bd()
    c.fill = PatternFill("solid", fgColor=C_LIGHT)
    c.alignment = Alignment(horizontal="right")
    _cell(ws, row, 3, "← change here to update all calculations", color="94A3B8")
    DR_ROW = row; row += 1

    _cell(ws, row, 1, "Time Horizon (years)", bold=True)
    c = ws.cell(row=row, column=2, value=data["time_horizon"])
    c.font = Font(name="Arial", bold=True, color=BLUE, size=10)
    c.number_format = "#,##0"; c.border = _bd()
    c.fill = PatternFill("solid", fgColor=C_LIGHT)
    c.alignment = Alignment(horizontal="right")
    _cell(ws, row, 3, "← change here to update all calculations", color="94A3B8")
    YR_ROW = row; row += 1

    _cell(ws, row, 1, "Currency", bold=True)
    _cell(ws, row, 2, cur)
    row += 2

    _sec(ws, row, 1, f"MEASURE INPUTS  [{cur}]  —  Blue cells are editable", span=n+3); row += 1

    _hdr(ws, row, 1, "Parameter", bg=C_ACCENT, sz=10)
    for ci, m in enumerate(measures, 2):
        _hdr(ws, row, ci, m["name"], bg=C_ACCENT, sz=10)
    _hdr(ws, row, n+2, "Notes", bg=C_ACCENT, sz=10)
    row += 1

    CAPEX_ROW = row
    _cell(ws, row, 1, "Capital Cost / CAPEX", bold=True)
    for ci, m in enumerate(measures, 2):
        c = ws.cell(row=row, column=ci, value=m["capex"])
        c.font = Font(name="Arial", bold=True, color=BLUE, size=10)
        c.number_format = "#,##0.0"; c.border = _bd()
        c.fill = PatternFill("solid", fgColor=C_LIGHT)
        c.alignment = Alignment(horizontal="right")
    _cell(ws, row, n+2, "One-time capital cost"); row += 1

    OPEX_ROW = row
    _cell(ws, row, 1, "Annual O&M Cost / OPEX", bold=True)
    for ci, m in enumerate(measures, 2):
        c = ws.cell(row=row, column=ci, value=m["annual_opex"])
        c.font = Font(name="Arial", bold=True, color=BLUE, size=10)
        c.number_format = "#,##0.0"; c.border = _bd()
        c.fill = PatternFill("solid", fgColor=C_LIGHT)
        c.alignment = Alignment(horizontal="right")
    _cell(ws, row, n+2, "Recurring annual cost"); row += 1

    BENEFIT_ROW = row
    _cell(ws, row, 1, "Annual Benefit", bold=True)
    for ci, m in enumerate(measures, 2):
        c = ws.cell(row=row, column=ci, value=m["annual_benefit"])
        c.font = Font(name="Arial", bold=True, color=BLUE, size=10)
        c.number_format = "#,##0.0"; c.border = _bd()
        c.fill = PatternFill("solid", fgColor=C_LIGHT)
        c.alignment = Alignment(horizontal="right")
    _cell(ws, row, n+2, "Annual monetised benefit"); row += 1

    LIFE_ROW = row
    _cell(ws, row, 1, "Measure Lifetime (years)", bold=True)
    for ci, m in enumerate(measures, 2):
        c = ws.cell(row=row, column=ci, value=m["lifetime_years"])
        c.font = Font(name="Arial", bold=True, color=BLUE, size=10)
        c.number_format = "#,##0"; c.border = _bd()
        c.fill = PatternFill("solid", fgColor=C_LIGHT)
        c.alignment = Alignment(horizontal="right")
    _cell(ws, row, n+2, "Used for PV calculations"); row += 1

    for label, key in [("Category", "category"), ("Feasibility", "feasibility"),
                       ("Uncertainty", "uncertainty"), ("Data Source", "data_source")]:
        _cell(ws, row, 1, label, bold=True)
        for ci, m in enumerate(measures, 2):
            _cell(ws, row, ci, m[key], align="center")
        row += 1

    _cell(ws, row, 1, "Benefit Types", bold=True)
    for ci, m in enumerate(measures, 2):
        _cell(ws, row, ci, ", ".join(m["benefit_types"]), align="center")
    row += 1

    _cell(ws, row, 1, "Co-benefits", bold=True)
    for ci, m in enumerate(measures, 2):
        _cell(ws, row, ci, m["co_benefits"])
    row += 2

    _sec(ws, row, 1, "KEY ASSUMPTIONS & LIMITATIONS", span=n+3); row += 1
    c = ws.cell(row=row, column=1, value=data.get("key_assumptions", ""))
    c.font = Font(name="Arial", size=9, color="475569")
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=row, start_column=1, end_row=row+2, end_column=n+3)
    ws.row_dimensions[row].height = 48; row += 3

    _cell(ws, row, 1, "Data Gaps:", bold=True)
    c = ws.cell(row=row, column=2, value=data.get("data_gaps", ""))
    c.font = Font(name="Arial", size=9, color="475569")
    c.alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=n+3); row += 2

    _sec(ws, row, 1, "COLOUR LEGEND", span=4); row += 1
    for txt, col in [("Blue = hardcoded input (editable)", BLUE),
                     ("Black = formula (do not overwrite)", BLACK),
                     ("Green = linked from another sheet (do not overwrite)", GREEN_LK)]:
        c = ws.cell(row=row, column=1, value=txt)
        c.font = Font(name="Arial", color=col, size=9)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        row += 1

    _widths(ws, {1: 30, **{i+2: 22 for i in range(n)}, n+2: 28})

    return {"dr": DR_ROW, "yr": YR_ROW,
            "capex": CAPEX_ROW, "opex": OPEX_ROW,
            "benefit": BENEFIT_ROW, "life": LIFE_ROW, "n": n}


# ── SHEET 2: CBA RESULTS ───────────────────────────────────────────────────────
def _results(wb, data, rm):
    ws = wb.create_sheet("CBA Results")
    ws.sheet_view.showGridLines = False
    measures = data["measures"]
    n = rm["n"]

    _hdr(ws, 1, 1, "COST-BENEFIT ANALYSIS — RESULTS", sz=13, span=n+3)
    _hdr(ws, 2, 1,
         f"All values in {data['currency']} {data['currency_unit']}  |  Change blue cells in Inputs sheet to update",
         bg=C_MID, fg="CBD5E1", bold=False, sz=9, span=n+3)

    row = 4
    _hdr(ws, row, 1, "Metric", bg=C_ACCENT, sz=10)
    for ci, m in enumerate(measures, 2):
        _hdr(ws, row, ci, m["name"], bg=C_ACCENT, sz=10)
    _hdr(ws, row, n+2, "Best", bg=C_ACCENT, sz=10)
    row += 1

    dr   = f"Inputs!$B${rm['dr']}"
    # Helper references
    def inp(key, ci):
        return f"Inputs!{get_column_letter(ci)}${rm[key]}"

    rr = {}  # track row numbers

    def add_row(label, key, f_fn, fmt, is_key=False):
        nonlocal row
        bg = C_LIGHT if is_key else None
        _cell(ws, row, 1, label, bold=is_key, bg=bg)
        for ci in range(2, n+2):
            f = f_fn(ci)
            c = ws.cell(row=row, column=ci, value=f)
            is_inp_link = f.startswith("=Inputs")
            c.font = Font(name="Arial", color=GREEN_LK if is_inp_link else BLACK,
                          bold=is_key, size=10)
            c.number_format = fmt; c.border = _bd()
            c.alignment = Alignment(horizontal="right", vertical="center")
            if bg: c.fill = PatternFill("solid", fgColor=C_LIGHT)
        rr[key] = row; row += 1

    add_row(f"CAPEX ({data['currency']}M)", "capex",
            lambda ci: f"=Inputs!{get_column_letter(ci)}${rm['capex']}", "#,##0.0")
    add_row(f"Annual Benefit ({data['currency']}M)", "ann_ben",
            lambda ci: f"=Inputs!{get_column_letter(ci)}${rm['benefit']}", "#,##0.0")
    add_row(f"Annual OPEX ({data['currency']}M)", "ann_opex",
            lambda ci: f"=Inputs!{get_column_letter(ci)}${rm['opex']}", "#,##0.0")

    def pv_ben_f(ci):
        b = inp("benefit", ci); l = inp("life", ci)
        return f"={b}/{dr}*(1-(1+{dr})^(-{l}))"

    def pv_cost_f(ci):
        k = inp("capex", ci); o = inp("opex", ci); l = inp("life", ci)
        return f"={k}+{o}/{dr}*(1-(1+{dr})^(-{l}))"

    add_row(f"PV of Benefits ({data['currency']}M)", "pv_ben", pv_ben_f, "#,##0.0")
    PVB = rr["pv_ben"]
    add_row(f"PV of Costs ({data['currency']}M)", "pv_cost", pv_cost_f, "#,##0.0")
    PVC = rr["pv_cost"]

    def npv_f(ci):
        col = get_column_letter(ci)
        return f"={col}{PVB}-{col}{PVC}"

    add_row(f"NPV ({data['currency']}M)", "npv", npv_f,
            '#,##0.0;(#,##0.0);"-"', is_key=True)
    NPV_ROW = rr["npv"]

    def bcr_f(ci):
        col = get_column_letter(ci)
        return f"=IF({col}{PVC}>0,{col}{PVB}/{col}{PVC},0)"

    add_row("BCR", "bcr", bcr_f, "0.00", is_key=True)
    BCR_ROW = rr["bcr"]

    def payback_f(ci):
        k = inp("capex", ci); b = inp("benefit", ci); o = inp("opex", ci)
        return f'=IF(({b}-{o})>0,{k}/({b}-{o}),"N/A")'

    add_row("Simple Payback (years)", "payback", payback_f, "0.0")

    def ann_net_f(ci):
        b = inp("benefit", ci); o = inp("opex", ci)
        return f"={b}-{o}"

    add_row(f"Annual Net Benefit ({data['currency']}M)", "ann_net", ann_net_f, "#,##0.0")

    # Best column formulas
    bcr_range = f"{get_column_letter(2)}{BCR_ROW}:{get_column_letter(n+1)}{BCR_ROW}"
    npv_range = f"{get_column_letter(2)}{NPV_ROW}:{get_column_letter(n+1)}{NPV_ROW}"

    for key_r, rng, fmt in [(BCR_ROW, bcr_range, "0.00"), (NPV_ROW, npv_range, "#,##0.0")]:
        c = ws.cell(row=key_r, column=n+2, value=f"=MAX({rng})")
        c.font = Font(name="Arial", color=C_GREEN, bold=True, size=10)
        c.number_format = fmt; c.border = _bd()
        c.fill = PatternFill("solid", fgColor=C_LIGHT)
        c.alignment = Alignment(horizontal="right")

    # Best measure name (above BCR row)
    name_range = f"{get_column_letter(2)}4:{get_column_letter(n+1)}4"
    c = ws.cell(row=4, column=n+2,
                value=f'=INDEX({name_range},MATCH(MAX({bcr_range}),{bcr_range},0))')
    c.font = Font(name="Arial", color=C_GREEN, bold=True, size=10)
    c.fill = PatternFill("solid", fgColor=C_DARK)
    c.alignment = Alignment(horizontal="center")

    # Ranking & Viability
    row += 1
    _sec(ws, row, 1, "RANKING & VIABILITY", span=n+2); row += 1

    _cell(ws, row, 1, "BCR Rank (1 = best)", bold=True)
    for ci in range(2, n+2):
        col = get_column_letter(ci)
        c = ws.cell(row=row, column=ci,
                    value=f"=RANK({col}{BCR_ROW},{bcr_range},0)")
        c.font = Font(name="Arial", color=BLACK, bold=True, size=10)
        c.number_format = "0"; c.border = _bd()
        c.alignment = Alignment(horizontal="center")
    row += 1

    _cell(ws, row, 1, "Viable? (BCR ≥ 1)", bold=True)
    for ci in range(2, n+2):
        col = get_column_letter(ci)
        c = ws.cell(row=row, column=ci,
                    value=f'=IF({col}{BCR_ROW}>=1,"✓ Yes","✗ No")')
        c.font = Font(name="Arial", color=BLACK, size=10)
        c.border = _bd()
        c.alignment = Alignment(horizontal="center")
    row += 2

    # Chart data area
    chart_r = row
    ws.cell(row=chart_r, column=1, value="Measure")
    ws.cell(row=chart_r, column=2, value="BCR")
    for mi, m in enumerate(measures, 1):
        ws.cell(row=chart_r+mi, column=1, value=m["name"])
        ws.cell(row=chart_r+mi, column=2,
                value=f"={get_column_letter(mi+1)}{BCR_ROW}")

    chart = BarChart()
    chart.type = "col"; chart.title = "BCR by Measure"
    chart.y_axis.title = "Benefit-Cost Ratio"
    chart.style = 10; chart.width = 18; chart.height = 12

    data_ref = Reference(ws, min_col=2, max_col=2, min_row=chart_r, max_row=chart_r+n)
    cats_ref = Reference(ws, min_col=1, min_row=chart_r+1, max_row=chart_r+n)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    ws.add_chart(chart, f"A{chart_r+n+3}")

    _widths(ws, {1: 32, **{i+2: 22 for i in range(n)}, n+2: 20})

    # Expose BCR_ROW for summary sheet
    rm["results_bcr_row"]   = BCR_ROW
    rm["results_capex_row"] = rr["capex"]
    rm["results_npv_row"]   = NPV_ROW


# ── SHEET 3: SENSITIVITY ──────────────────────────────────────────────────────
def _sensitivity(wb, data, rm):
    ws = wb.create_sheet("Sensitivity")
    ws.sheet_view.showGridLines = False
    measures = data["measures"]
    n = rm["n"]
    sens_vars = data.get("sensitivity_vars", [])

    _hdr(ws, 1, 1, "SENSITIVITY ANALYSIS", sz=13, span=n+5)
    _hdr(ws, 2, 1,
         "Edit Low / Base / High values (blue cells). BCR table below recalculates automatically via formulas.",
         bg=C_MID, fg="CBD5E1", bold=False, sz=9, span=n+5)

    row = 4
    _sec(ws, row, 1, "PARAMETER RANGES  (edit blue cells)", span=5); row += 1
    for ci, h in enumerate(["Variable", "Low", "Base", "High", "Unit"], 1):
        _hdr(ws, row, ci, h, bg=C_ACCENT, sz=10)
    row += 1

    sv_rows = {}
    for sv in sens_vars:
        _cell(ws, row, 1, sv["name"], bold=True)
        for col_i, val in enumerate([sv["low"], sv["base"], sv["high"]], 2):
            c = ws.cell(row=row, column=col_i, value=val)
            c.font = Font(name="Arial", bold=True, color=BLUE, size=10)
            c.number_format = "0.00%" if sv["unit"] == "%" else "0.00"
            c.border = _bd()
            c.fill = PatternFill("solid", fgColor=C_LIGHT)
            c.alignment = Alignment(horizontal="right")
        _cell(ws, row, 5, sv["unit"])
        sv_rows[sv["name"]] = {
            "low":  f"$B${row}",
            "base": f"$C${row}",
            "high": f"$D${row}"
        }
        row += 1

    row += 1
    _sec(ws, row, 1, "BCR SENSITIVITY TABLE  (formulas — updates automatically)", span=n+3); row += 1
    _hdr(ws, row, 1, "Variable", bg=C_ACCENT, sz=10)
    _hdr(ws, row, 2, "Scenario", bg=C_ACCENT, sz=10)
    for ci, m in enumerate(measures, 3):
        _hdr(ws, row, ci, m["name"], bg=C_ACCENT, sz=10)
    row += 1

    dr_inp = f"Inputs!$B${rm['dr']}"

    for sv in sens_vars:
        sv_name = sv["name"]
        sv_r = sv_rows.get(sv_name, {})

        for scenario, col_key, color in [("Low", "low", C_RED),
                                          ("Base", "base", BLACK),
                                          ("High", "high", C_GREEN)]:
            param_ref = sv_r.get(col_key, "$C$7")
            bg = C_LIGHT if scenario == "Base" else None

            _cell(ws, row, 1, sv_name if scenario == "Low" else "", bold=(scenario == "Low"))
            _cell(ws, row, 2, scenario, color=color, bg=bg)

            for mi in range(n):
                ci = mi + 2  # column in Inputs sheet (measures start at col 2)
                capex = f"Inputs!{get_column_letter(ci)}${rm['capex']}"
                opex  = f"Inputs!{get_column_letter(ci)}${rm['opex']}"
                ben   = f"Inputs!{get_column_letter(ci)}${rm['benefit']}"
                life  = f"Inputs!{get_column_letter(ci)}${rm['life']}"

                if sv_name == "Discount Rate":
                    dr_use = param_ref
                    ben_term = ben
                elif sv_name == "Annual Benefit":
                    dr_use = dr_inp
                    ben_term = f"({ben}*{param_ref})"
                else:
                    dr_use = dr_inp
                    ben_term = ben

                pv_ben  = f"({ben_term}/{dr_use}*(1-(1+{dr_use})^(-{life})))"
                pv_cost = f"({capex}+{opex}/{dr_use}*(1-(1+{dr_use})^(-{life})))"
                formula = f"=IF({dr_use}>0,IF({pv_cost}>0,{pv_ben}/{pv_cost},0),0)"

                c = ws.cell(row=row, column=mi+3, value=formula)
                c.font = Font(name="Arial", color=color, bold=(scenario == "Base"), size=10)
                c.number_format = "0.00"; c.border = _bd()
                c.alignment = Alignment(horizontal="right")
                if bg: c.fill = PatternFill("solid", fgColor=C_LIGHT)
            row += 1
        row += 1  # blank between variables

    row += 1
    _sec(ws, row, 1, "INSTRUCTIONS", span=6); row += 1
    for inst in [
        "1. Edit Low / Base / High values in the table above (blue cells only)",
        "2. BCR values in the table update automatically — no manual recalculation needed",
        "3. 'Discount Rate': enter as decimal (e.g. 0.02 = 2%)",
        "4. 'Annual Benefit' multiplier: 1.0 = base, 0.5 = 50% of base benefits, 1.5 = 150%",
        "5. Add more sensitivity variables by editing the app and rerunning the analysis",
    ]:
        c = ws.cell(row=row, column=1, value=inst)
        c.font = Font(name="Arial", size=9, color="475569")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        row += 1

    _widths(ws, {1: 26, 2: 12, 3: 14, 4: 14, 5: 12, **{i+5: 22 for i in range(n)}})


# ── SHEET 4: SUMMARY ──────────────────────────────────────────────────────────
def _summary(wb, data, rm):
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False
    measures = data["measures"]
    n = rm["n"]
    BCR_ROW   = rm["results_bcr_row"]
    CAPEX_ROW = rm["results_capex_row"]
    NPV_ROW   = rm["results_npv_row"]

    _hdr(ws, 1, 1, "EXECUTIVE SUMMARY", sz=13, span=5)
    _hdr(ws, 2, 1, data["problem_title"], bg=C_MID, fg="CBD5E1", sz=11, span=5)

    row = 4
    _sec(ws, row, 1, "PROBLEM OVERVIEW", span=5); row += 1
    c = ws.cell(row=row, column=1, value=data["problem_summary"])
    c.font = Font(name="Arial", size=10)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=row, start_column=1, end_row=row+1, end_column=5)
    ws.row_dimensions[row].height = 40; row += 3

    _sec(ws, row, 1, "MEASURES AT A GLANCE  (values link live from CBA Results)", span=5); row += 1
    for ci, h in enumerate(["Measure", "Category", f"CAPEX ({data['currency']}M)",
                             "BCR (base)", "Viability"], 1):
        _hdr(ws, row, ci, h, bg=C_ACCENT, sz=10)
    row += 1

    for mi, m in enumerate(measures):
        ci = mi + 2  # Excel column in Results sheet
        col = get_column_letter(ci)

        _cell(ws, row, 1, m["name"], bold=True)
        _cell(ws, row, 2, m["category"], align="center")

        c = ws.cell(row=row, column=3, value=f"='CBA Results'!{col}{CAPEX_ROW}")
        c.font = Font(name="Arial", color=GREEN_LK, size=10)
        c.number_format = "#,##0.0"; c.border = _bd()
        c.alignment = Alignment(horizontal="right")

        c = ws.cell(row=row, column=4, value=f"='CBA Results'!{col}{BCR_ROW}")
        c.font = Font(name="Arial", color=GREEN_LK, bold=True, size=10)
        c.number_format = "0.00"; c.border = _bd()
        c.fill = PatternFill("solid", fgColor=C_LIGHT)
        c.alignment = Alignment(horizontal="center")

        bcr_here = f"D{row}"
        c = ws.cell(row=row, column=5,
                    value=f'=IF({bcr_here}>=1.5,"✓ Recommended",IF({bcr_here}>=1.0,"○ Consider","✗ Review"))')
        c.font = Font(name="Arial", color=BLACK, size=10)
        c.border = _bd(); c.alignment = Alignment(horizontal="center")
        row += 1

    row += 1
    _sec(ws, row, 1, "HOW TO USE THIS WORKBOOK", span=5); row += 1
    for step in [
        "1.  INPUTS sheet — Edit blue cells (CAPEX, OPEX, Benefits, Discount Rate, Time Horizon)",
        "2.  CBA RESULTS sheet — NPV and BCR recalculate automatically from Inputs",
        "3.  SENSITIVITY sheet — Edit Low/Base/High ranges; BCR table updates automatically",
        "4.  SUMMARY sheet (this sheet) — BCR and CAPEX values link live from CBA Results",
        "5.  Never overwrite black or green cells — they contain formulas",
    ]:
        c = ws.cell(row=row, column=1, value=step)
        c.font = Font(name="Arial", size=10)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        row += 1

    _widths(ws, {1: 30, 2: 18, 3: 18, 4: 16, 5: 18})
